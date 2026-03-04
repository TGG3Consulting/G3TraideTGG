# -*- coding: utf-8 -*-
"""
Genetic Algorithm for Feature Selection.

Professional implementation with:
- Binary chromosome encoding
- Tournament selection
- Two-point crossover
- Swap mutation (maintains exact feature count)
- Elitism
- Early stopping
- Checkpointing
- Comprehensive logging
- Full results export to Excel
"""

import json
import pickle
import random
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional
import signal
import sys
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.preprocessing import StandardScaler

from .config import GAConfig, FeaturePool

warnings.filterwarnings('ignore')


@dataclass
class Individual:
    """Single individual in the population (chromosome)."""
    genes: List[int]  # Binary: 1 = feature selected, 0 = not selected
    fitness: float = 0.0
    metrics: Dict[str, float] = field(default_factory=dict)
    features: List[str] = field(default_factory=list)

    def __hash__(self):
        return hash(tuple(self.genes))

    def __eq__(self, other):
        return self.genes == other.genes


@dataclass
class GenerationStats:
    """Statistics for one generation."""
    generation: int
    best_fitness: float
    avg_fitness: float
    worst_fitness: float
    best_metrics: Dict[str, float]
    best_features: List[str]
    diversity: float  # Unique individuals / population size
    timestamp: str


@dataclass
class ModelMetrics:
    """Metrics for a single model."""
    name: str
    f1: float = 0.0
    precision: float = 0.0
    recall: float = 0.0
    auc: float = 0.0
    accuracy: float = 0.0
    # Threshold analysis
    thresh_050_prec: float = 0.0
    thresh_050_rec: float = 0.0
    thresh_050_wr: float = 0.0
    thresh_050_pnl: float = 0.0
    thresh_050_cnt: int = 0
    thresh_070_prec: float = 0.0
    thresh_070_rec: float = 0.0
    thresh_070_wr: float = 0.0
    thresh_070_pnl: float = 0.0
    thresh_070_cnt: int = 0
    thresh_090_prec: float = 0.0
    thresh_090_rec: float = 0.0
    thresh_090_wr: float = 0.0
    thresh_090_pnl: float = 0.0
    thresh_090_cnt: int = 0


@dataclass
class EvaluationRecord:
    """Record of a single evaluation for Excel export - ALL 5 MODELS."""
    eval_id: int
    generation: int
    features: List[str]
    # Aggregate metrics (average across models)
    avg_f1: float = 0.0
    avg_precision: float = 0.0
    avg_recall: float = 0.0
    avg_auc: float = 0.0
    # Best model
    best_model: str = ""
    best_f1: float = 0.0
    # Individual model metrics
    lgb: ModelMetrics = field(default_factory=lambda: ModelMetrics("LightGBM"))
    xgb: ModelMetrics = field(default_factory=lambda: ModelMetrics("XGBoost"))
    cb: ModelMetrics = field(default_factory=lambda: ModelMetrics("CatBoost"))
    rf: ModelMetrics = field(default_factory=lambda: ModelMetrics("RandomForest"))
    lr: ModelMetrics = field(default_factory=lambda: ModelMetrics("LogisticRegression"))


@dataclass
class GAResult:
    """Final result of GA optimization."""
    best_individual: Individual
    best_features: List[str]
    best_metrics: Dict[str, float]
    generation_history: List[GenerationStats]
    all_evaluations: List[EvaluationRecord]
    total_evaluations: int
    total_time_seconds: float
    config: Dict[str, Any]


class GeneticFeatureSelector:
    """
    Genetic Algorithm for optimal feature subset selection.

    Usage:
        selector = GeneticFeatureSelector(config, feature_pool)
        result = selector.run(df, target_column="label_win")
    """

    def __init__(self, config: GAConfig, feature_pool: FeaturePool):
        self.config = config
        self.feature_pool = feature_pool
        self.all_features: List[str] = []
        self.n_features: int = 0

        # State
        self.population: List[Individual] = []
        self.generation: int = 0
        self.best_ever: Optional[Individual] = None
        self.generations_without_improvement: int = 0
        self.history: List[GenerationStats] = []
        self.total_evaluations: int = 0
        self.all_evaluations: List[EvaluationRecord] = []  # ALL evaluated combinations
        self._evaluation_cache: Dict[tuple, Dict] = {}  # Cache: genes tuple -> metrics

        # Data (set during run)
        self._df: Optional[pd.DataFrame] = None
        self._target_column: str = ""
        self._X_train: Optional[pd.DataFrame] = None
        self._X_test: Optional[pd.DataFrame] = None
        self._y_train: Optional[np.ndarray] = None
        self._y_test: Optional[np.ndarray] = None

        # Random state
        if config.random_seed is not None:
            random.seed(config.random_seed)
            np.random.seed(config.random_seed)

        # Validate config
        if config.elite_count >= config.population_size:
            raise ValueError(f"elite_count ({config.elite_count}) must be < population_size ({config.population_size})")
        if config.n_features_to_select < 1:
            raise ValueError("n_features_to_select must be >= 1")
        if config.tournament_size > config.population_size:
            raise ValueError("tournament_size cannot exceed population_size")

        # Output directory
        self.config.output_dir.mkdir(parents=True, exist_ok=True)

    def run(self, df: pd.DataFrame, target_column: str = "label_win") -> GAResult:
        """
        Run the genetic algorithm.

        Args:
            df: DataFrame with features and target
            target_column: Name of the target column

        Returns:
            GAResult with best features and metrics
        """
        start_time = time.time()
        self._start_time = start_time
        self._interrupted = False

        # Setup Ctrl+C handler to save checkpoint
        def signal_handler(signum, frame):
            print("\n\nInterrupted! Saving checkpoint...")
            self._interrupted = True
            self._save_checkpoint()
            print(f"Checkpoint saved. Resume with --resume flag.")
            sys.exit(0)

        original_handler = signal.signal(signal.SIGINT, signal_handler)

        try:
            # Setup
            self._setup_data(df, target_column)
            self._initialize_population()

            if self.config.verbose:
                self._print_header()

            import gc

            # Evolution loop
            while self.generation < self.config.n_generations:
                self._evolve_generation()

                if self.config.verbose and self.generation % self.config.log_every_n_generations == 0:
                    self._print_generation_stats()

                # Checkpoint
                if self.generation % self.config.save_every_n_generations == 0:
                    self._save_checkpoint()

                # Periodic memory cleanup (every 10 generations)
                if self.generation % 10 == 0:
                    gc.collect()

                # Early stopping
                if self.generations_without_improvement >= self.config.early_stopping_rounds:
                    if self.config.verbose:
                        print(f"\nEarly stopping at generation {self.generation} "
                              f"(no improvement for {self.config.early_stopping_rounds} generations)")
                    break

                self.generation += 1

            # Final result
            total_time = time.time() - start_time
            result = self._create_result(total_time)

            # Save final results
            self._save_results(result)

            if self.config.verbose:
                self._print_final_results(result)

            return result

        finally:
            # Restore original signal handler
            signal.signal(signal.SIGINT, original_handler)

    def _setup_data(self, df: pd.DataFrame, target_column: str):
        """Prepare data for training."""
        self._df = df.copy()
        self._target_column = target_column

        # Validate features
        self.all_features = self.feature_pool.validate_features(list(df.columns))
        self.n_features = len(self.all_features)

        if self.n_features < self.config.n_features_to_select:
            raise ValueError(
                f"Not enough features: {self.n_features} available, "
                f"{self.config.n_features_to_select} requested"
            )

        # Add direction_num if Direction column exists
        if "Direction" in df.columns:
            self._df["direction_num"] = (df["Direction"] == "LONG").astype(int)

        # Temporal train/test split
        split_idx = int(len(df) * self.config.train_test_split)
        train_df = self._df.iloc[:split_idx]
        test_df = self._df.iloc[split_idx:]

        self._y_train = train_df[target_column].values
        self._y_test = test_df[target_column].values
        self._train_df = train_df
        self._test_df = test_df

        if self.config.verbose:
            print(f"Features available: {self.n_features}")
            print(f"Features to select: {self.config.n_features_to_select}")
            print(f"Train size: {len(train_df)}, Test size: {len(test_df)}")
            print(f"Target distribution - Train: {self._y_train.mean():.1%}, Test: {self._y_test.mean():.1%}")

    def _initialize_population(self):
        """Create initial random population."""
        self.population = []

        for _ in range(self.config.population_size):
            individual = self._create_random_individual()
            self.population.append(individual)

        # Evaluate initial population
        self._evaluate_population()

        # Sort by fitness
        self.population.sort(key=lambda x: x.fitness, reverse=True)

        # Set best ever
        self.best_ever = self.population[0]

        # Record initial stats
        self._record_generation_stats()

    def _create_random_individual(self) -> Individual:
        """Create a random individual with exactly n_features_to_select features."""
        genes = [0] * self.n_features

        # Randomly select n features
        selected_indices = random.sample(range(self.n_features), self.config.n_features_to_select)
        for idx in selected_indices:
            genes[idx] = 1

        return Individual(genes=genes)

    def _evaluate_population(self):
        """Evaluate fitness of all individuals in population."""
        for individual in self.population:
            if individual.fitness == 0.0:  # Not yet evaluated
                self._evaluate_individual(individual)
                self.total_evaluations += 1

    def _evaluate_individual(self, individual: Individual):
        """
        Evaluate a single individual with ALL 5 MODELS.

        Trains all models on selected features and computes fitness metrics.
        Includes full threshold analysis and saves to all_evaluations.
        Uses caching to avoid re-evaluating duplicate combinations.
        """
        # Check cache first
        genes_key = tuple(individual.genes)
        if genes_key in self._evaluation_cache:
            cached = self._evaluation_cache[genes_key]
            individual.metrics = cached["metrics"]
            individual.fitness = cached["fitness"]
            individual.features = cached["features"]
            return  # Skip re-evaluation

        # Get selected features
        selected_features = [
            self.all_features[i]
            for i, gene in enumerate(individual.genes)
            if gene == 1
        ]
        individual.features = selected_features

        # Add direction_num if available
        feature_cols = selected_features.copy()
        if "direction_num" in self._df.columns:
            feature_cols.append("direction_num")

        # Prepare data
        X_train = self._train_df[feature_cols].values
        X_test = self._test_df[feature_cols].values

        # Handle NaN
        X_train = np.nan_to_num(X_train, nan=0.0)
        X_test = np.nan_to_num(X_test, nan=0.0)

        # Train and evaluate ALL 5 models
        all_model_results = self._train_all_models(X_train, X_test, self._y_train, self._y_test)

        # Calculate aggregate metrics
        f1_scores = [r["metrics"]["f1"] for r in all_model_results.values()]
        prec_scores = [r["metrics"]["precision"] for r in all_model_results.values()]
        rec_scores = [r["metrics"]["recall"] for r in all_model_results.values()]
        auc_scores = [r["metrics"]["auc"] for r in all_model_results.values()]

        avg_f1 = np.mean(f1_scores)
        avg_precision = np.mean(prec_scores)
        avg_recall = np.mean(rec_scores)
        avg_auc = np.mean(auc_scores)

        # Find best model
        best_model_name = max(all_model_results.keys(), key=lambda k: all_model_results[k]["metrics"]["f1"])
        best_f1 = all_model_results[best_model_name]["metrics"]["f1"]

        # Set individual fitness (use average F1 for robustness)
        individual.metrics = {
            "f1": avg_f1,
            "precision": avg_precision,
            "recall": avg_recall,
            "auc": avg_auc,
            "best_model": best_model_name,
            "best_f1": best_f1,
        }
        individual.fitness = avg_f1  # Use average for GA selection

        # Create evaluation record with all model results
        record = EvaluationRecord(
            eval_id=self.total_evaluations,
            generation=self.generation,
            features=selected_features.copy(),
            avg_f1=avg_f1,
            avg_precision=avg_precision,
            avg_recall=avg_recall,
            avg_auc=avg_auc,
            best_model=best_model_name,
            best_f1=best_f1,
        )

        # Fill in individual model metrics
        model_map = {"lightgbm": "lgb", "xgboost": "xgb", "catboost": "cb", "rf": "rf", "lr": "lr"}
        for model_name, results in all_model_results.items():
            model_attr = model_map[model_name]
            model_metrics = getattr(record, model_attr)
            model_metrics.f1 = results["metrics"]["f1"]
            model_metrics.precision = results["metrics"]["precision"]
            model_metrics.recall = results["metrics"]["recall"]
            model_metrics.auc = results["metrics"]["auc"]
            model_metrics.accuracy = results["metrics"]["accuracy"]

            # Threshold results
            for thresh in [0.50, 0.70, 0.90]:
                thresh_key = f"thresh_{int(thresh*100):03d}"
                thresh_data = results["thresholds"].get(thresh, {})
                setattr(model_metrics, f"{thresh_key}_prec", thresh_data.get("precision", 0.0))
                setattr(model_metrics, f"{thresh_key}_rec", thresh_data.get("recall", 0.0))
                setattr(model_metrics, f"{thresh_key}_wr", thresh_data.get("win_rate", 0.0))
                setattr(model_metrics, f"{thresh_key}_pnl", thresh_data.get("avg_pnl", 0.0))
                setattr(model_metrics, f"{thresh_key}_cnt", thresh_data.get("count", 0))

        self.all_evaluations.append(record)

        # Save to cache
        self._evaluation_cache[genes_key] = {
            "metrics": individual.metrics,
            "fitness": individual.fitness,
            "features": individual.features,
        }

    def _train_all_models(
        self,
        X_train: np.ndarray,
        X_test: np.ndarray,
        y_train: np.ndarray,
        y_test: np.ndarray
    ) -> Dict[str, Dict]:
        """
        Train ALL 5 models and return their metrics with threshold analysis.

        Returns:
            Dict with model_name -> {"metrics": {...}, "thresholds": {...}}
        """
        from sklearn.metrics import (
            accuracy_score, precision_score, recall_score,
            f1_score, roc_auc_score
        )
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.linear_model import LogisticRegression
        import lightgbm as lgb
        import xgboost as xgb
        from catboost import CatBoostClassifier

        # Define all models
        models = {
            "lightgbm": lgb.LGBMClassifier(
                n_estimators=100, max_depth=6, learning_rate=0.1,
                num_leaves=31, verbose=-1, random_state=self.config.random_seed, n_jobs=-1
            ),
            "xgboost": xgb.XGBClassifier(
                n_estimators=100, max_depth=6, learning_rate=0.1,
                verbosity=0, random_state=self.config.random_seed, n_jobs=-1
            ),
            "catboost": CatBoostClassifier(
                iterations=100, depth=6, learning_rate=0.1,
                verbose=False, random_state=self.config.random_seed,
                allow_writing_files=False  # No temp files
            ),
            "rf": RandomForestClassifier(
                n_estimators=100, max_depth=10,
                random_state=self.config.random_seed, n_jobs=-1
            ),
            "lr": LogisticRegression(
                max_iter=1000, random_state=self.config.random_seed,
                solver='lbfgs', n_jobs=-1, class_weight='balanced'  # Handle imbalance
            ),
        }

        # Scale data for LogisticRegression (must be before lr_X_train assignment)
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)

        # LR uses scaled data
        lr_X_train, lr_X_test = X_train_scaled, X_test_scaled

        # Handle NaN in PnL
        if "Net %" in self._test_df.columns:
            test_pnl = self._test_df["Net %"].values.copy()
            test_pnl = np.nan_to_num(test_pnl, nan=0.0)
        else:
            test_pnl = None

        results = {}

        import gc

        for model_name, model in models.items():
            try:
                # Use scaled data for LR
                if model_name == "lr":
                    X_tr, X_te = lr_X_train, lr_X_test
                else:
                    X_tr, X_te = X_train, X_test

                # Train
                model.fit(X_tr, y_train)

                # Predict
                y_pred = model.predict(X_te)
                y_proba = model.predict_proba(X_te)[:, 1]

                # Base metrics
                metrics = {
                    "accuracy": accuracy_score(y_test, y_pred),
                    "precision": precision_score(y_test, y_pred, zero_division=0),
                    "recall": recall_score(y_test, y_pred, zero_division=0),
                    "f1": f1_score(y_test, y_pred, zero_division=0),
                    "auc": roc_auc_score(y_test, y_proba) if len(np.unique(y_test)) > 1 else 0.5,
                }

                # Threshold analysis (0.50, 0.70, 0.90)
                thresholds_data = {}
                for thresh in [0.50, 0.70, 0.90]:
                    y_pred_thresh = (y_proba >= thresh).astype(int)
                    pred_positive = y_pred_thresh == 1

                    if pred_positive.sum() > 0:
                        tp = ((y_pred_thresh == 1) & (y_test == 1)).sum()
                        fp = ((y_pred_thresh == 1) & (y_test == 0)).sum()
                        fn = ((y_pred_thresh == 0) & (y_test == 1)).sum()

                        thresholds_data[thresh] = {
                            "precision": tp / (tp + fp) if (tp + fp) > 0 else 0.0,
                            "recall": tp / (tp + fn) if (tp + fn) > 0 else 0.0,
                            "win_rate": y_test[pred_positive].mean(),
                            "avg_pnl": test_pnl[pred_positive].mean() if test_pnl is not None else 0.0,
                            "count": int(pred_positive.sum()),
                        }
                    else:
                        thresholds_data[thresh] = {
                            "precision": 0.0, "recall": 0.0, "win_rate": 0.0, "avg_pnl": 0.0, "count": 0
                        }

                results[model_name] = {"metrics": metrics, "thresholds": thresholds_data}

            except Exception as e:
                # If model fails, use zeros
                results[model_name] = {
                    "metrics": {"accuracy": 0, "precision": 0, "recall": 0, "f1": 0, "auc": 0.5},
                    "thresholds": {t: {"precision": 0, "recall": 0, "win_rate": 0, "avg_pnl": 0, "count": 0} for t in [0.5, 0.7, 0.9]}
                }

            finally:
                # Explicit cleanup to prevent memory leaks
                del model

        # Force garbage collection after all models
        del models
        gc.collect()

        return results

    def _evolve_generation(self):
        """Perform one generation of evolution."""
        new_population = []

        # Elitism: keep best individuals
        elite = self.population[:self.config.elite_count]
        new_population.extend(elite)

        # Generate offspring
        while len(new_population) < self.config.population_size:
            # Selection
            parent1 = self._tournament_select()
            parent2 = self._tournament_select()

            # Crossover
            if random.random() < self.config.crossover_rate:
                child1, child2 = self._crossover(parent1, parent2)
            else:
                child1 = Individual(genes=parent1.genes.copy())
                child2 = Individual(genes=parent2.genes.copy())

            # Mutation
            if random.random() < self.config.mutation_rate:
                self._mutate(child1)
            if random.random() < self.config.mutation_rate:
                self._mutate(child2)

            # Repair if needed (ensure exactly n_features_to_select)
            self._repair(child1)
            self._repair(child2)

            new_population.append(child1)
            if len(new_population) < self.config.population_size:
                new_population.append(child2)

        self.population = new_population

        # Evaluate new individuals
        self._evaluate_population()

        # Sort by fitness
        self.population.sort(key=lambda x: x.fitness, reverse=True)

        # Update best ever
        if self.population[0].fitness > self.best_ever.fitness:
            self.best_ever = self.population[0]
            self.generations_without_improvement = 0
        else:
            self.generations_without_improvement += 1

        # Record stats
        self._record_generation_stats()

    def _tournament_select(self) -> Individual:
        """Select individual using tournament selection."""
        tournament = random.sample(self.population, self.config.tournament_size)
        return max(tournament, key=lambda x: x.fitness)

    def _crossover(self, parent1: Individual, parent2: Individual) -> Tuple[Individual, Individual]:
        """Perform crossover between two parents."""
        genes1 = parent1.genes.copy()
        genes2 = parent2.genes.copy()

        if self.config.crossover_type == "single_point":
            point = random.randint(1, self.n_features - 1)
            child1_genes = genes1[:point] + genes2[point:]
            child2_genes = genes2[:point] + genes1[point:]

        elif self.config.crossover_type == "two_point":
            point1 = random.randint(1, self.n_features - 2)
            point2 = random.randint(point1 + 1, self.n_features - 1)
            child1_genes = genes1[:point1] + genes2[point1:point2] + genes1[point2:]
            child2_genes = genes2[:point1] + genes1[point1:point2] + genes2[point2:]

        else:  # uniform
            child1_genes = []
            child2_genes = []
            for g1, g2 in zip(genes1, genes2):
                if random.random() < 0.5:
                    child1_genes.append(g1)
                    child2_genes.append(g2)
                else:
                    child1_genes.append(g2)
                    child2_genes.append(g1)

        return Individual(genes=child1_genes), Individual(genes=child2_genes)

    def _mutate(self, individual: Individual):
        """
        Mutate individual using swap mutation.

        Swap mutation: randomly swap a selected feature with an unselected one.
        This maintains exactly n_features_to_select features.
        """
        selected_indices = [i for i, g in enumerate(individual.genes) if g == 1]
        unselected_indices = [i for i, g in enumerate(individual.genes) if g == 0]

        if selected_indices and unselected_indices:
            # Swap one selected with one unselected
            idx_to_remove = random.choice(selected_indices)
            idx_to_add = random.choice(unselected_indices)

            individual.genes[idx_to_remove] = 0
            individual.genes[idx_to_add] = 1

    def _repair(self, individual: Individual):
        """Repair individual to have exactly n_features_to_select features."""
        n_selected = sum(individual.genes)
        target = self.config.n_features_to_select

        if n_selected == target:
            return

        selected_indices = [i for i, g in enumerate(individual.genes) if g == 1]
        unselected_indices = [i for i, g in enumerate(individual.genes) if g == 0]

        if n_selected > target:
            # Remove random features
            to_remove = random.sample(selected_indices, n_selected - target)
            for idx in to_remove:
                individual.genes[idx] = 0

        elif n_selected < target:
            # Add random features
            to_add = random.sample(unselected_indices, target - n_selected)
            for idx in to_add:
                individual.genes[idx] = 1

    def _record_generation_stats(self):
        """Record statistics for current generation."""
        fitnesses = [ind.fitness for ind in self.population]

        # Calculate diversity (unique individuals)
        unique_count = len(set(tuple(ind.genes) for ind in self.population))
        diversity = unique_count / len(self.population)

        stats = GenerationStats(
            generation=self.generation,
            best_fitness=max(fitnesses),
            avg_fitness=np.mean(fitnesses),
            worst_fitness=min(fitnesses),
            best_metrics=self.population[0].metrics.copy(),
            best_features=self.population[0].features.copy(),
            diversity=diversity,
            timestamp=datetime.now().isoformat()
        )

        self.history.append(stats)

    def _save_checkpoint(self):
        """Save checkpoint for resuming."""
        checkpoint = {
            "generation": self.generation,
            "population": self.population,
            "best_ever": self.best_ever,
            "history": self.history,
            "total_evaluations": self.total_evaluations,
            "generations_without_improvement": self.generations_without_improvement,
            "all_features": self.all_features,
            "all_evaluations": self.all_evaluations,
            "evaluation_cache": self._evaluation_cache,
        }

        checkpoint_path = self.config.output_dir / self.config.checkpoint_file
        with open(checkpoint_path, "wb") as f:
            pickle.dump(checkpoint, f)

    def load_checkpoint(self, checkpoint_path: Path) -> bool:
        """Load checkpoint to resume evolution."""
        if not checkpoint_path.exists():
            return False

        with open(checkpoint_path, "rb") as f:
            checkpoint = pickle.load(f)

        self.generation = checkpoint["generation"]
        self.population = checkpoint["population"]
        self.best_ever = checkpoint["best_ever"]
        self.history = checkpoint["history"]
        self.total_evaluations = checkpoint["total_evaluations"]
        self.generations_without_improvement = checkpoint["generations_without_improvement"]
        self.all_features = checkpoint["all_features"]
        self.all_evaluations = checkpoint.get("all_evaluations", [])
        self._evaluation_cache = checkpoint.get("evaluation_cache", {})

        return True

    def _create_result(self, total_time: float) -> GAResult:
        """Create final result object."""
        return GAResult(
            best_individual=self.best_ever,
            best_features=self.best_ever.features,
            best_metrics=self.best_ever.metrics,
            generation_history=self.history,
            all_evaluations=self.all_evaluations,
            total_evaluations=self.total_evaluations,
            total_time_seconds=total_time,
            config=asdict(self.config)
        )

    def _save_results(self, result: GAResult):
        """Save final results to JSON and Excel."""
        # Save JSON summary
        results_path = self.config.output_dir / self.config.results_file
        results_dict = {
            "best_features": result.best_features,
            "best_metrics": result.best_metrics,
            "total_evaluations": result.total_evaluations,
            "total_time_seconds": result.total_time_seconds,
            "final_generation": self.generation,
            "config": {k: str(v) if isinstance(v, Path) else v for k, v in result.config.items()},
        }

        with open(results_path, "w", encoding="utf-8") as f:
            json.dump(results_dict, f, indent=2, ensure_ascii=False)

        # Save best features as simple list
        features_path = self.config.output_dir / "best_features.txt"
        with open(features_path, "w") as f:
            f.write("\n".join(result.best_features))

        # Save ALL evaluations to Excel
        self._save_evaluations_excel(result)

    def _save_evaluations_excel(self, result: GAResult):
        """Save all evaluations to a formatted Excel file with ALL 5 MODELS."""
        excel_path = self.config.output_dir / "ga_all_results.xlsx"

        wb = Workbook()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        model_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ===== Sheet 1: Summary (Aggregate) =====
        ws1 = wb.active
        ws1.title = "Summary"

        headers1 = [
            "ID", "Gen", "Avg_F1", "Avg_Prec", "Avg_Rec", "Avg_AUC",
            "Best_Model", "Best_F1",
            "LGB_F1", "XGB_F1", "CB_F1", "RF_F1", "LR_F1",
            "Features"
        ]

        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for row_idx, record in enumerate(result.all_evaluations, 2):
            row_data = [
                record.eval_id, record.generation,
                record.avg_f1, record.avg_precision, record.avg_recall, record.avg_auc,
                record.best_model, record.best_f1,
                record.lgb.f1, record.xgb.f1, record.cb.f1, record.rf.f1, record.lr.f1,
                ", ".join(record.features),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '0.0000'

        # Adjust column widths
        for col in ['A', 'B']:
            ws1.column_dimensions[col].width = 5
        for col in ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws1.column_dimensions[col].width = 9
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['N'].width = 80

        # ===== Sheet 2: LightGBM Details =====
        self._write_model_sheet(wb, "LightGBM", result.all_evaluations, "lgb", header_font, header_fill, border)

        # ===== Sheet 3: XGBoost Details =====
        self._write_model_sheet(wb, "XGBoost", result.all_evaluations, "xgb", header_font, header_fill, border)

        # ===== Sheet 4: CatBoost Details =====
        self._write_model_sheet(wb, "CatBoost", result.all_evaluations, "cb", header_font, header_fill, border)

        # ===== Sheet 5: RandomForest Details =====
        self._write_model_sheet(wb, "RandomForest", result.all_evaluations, "rf", header_font, header_fill, border)

        # ===== Sheet 6: LogisticRegression Details =====
        self._write_model_sheet(wb, "LogisticReg", result.all_evaluations, "lr", header_font, header_fill, border)

        # ===== Sheet 7: Top 20 by Avg F1 =====
        ws_top = wb.create_sheet("Top20_AvgF1")
        sorted_records = sorted(result.all_evaluations, key=lambda x: x.avg_f1, reverse=True)[:20]
        self._write_top_summary_sheet(ws_top, sorted_records, headers1, header_font, header_fill, border)

        # ===== Sheet 8: Top 20 by Best Single Model F1 =====
        ws_best = wb.create_sheet("Top20_BestF1")
        sorted_best = sorted(result.all_evaluations, key=lambda x: x.best_f1, reverse=True)[:20]
        self._write_top_summary_sheet(ws_best, sorted_best, headers1, header_font, header_fill, border)

        # ===== Sheet 9: Config & Stats =====
        ws_info = wb.create_sheet("Info")
        info_data = [
            ["GA Feature Selection Results", ""],
            ["", ""],
            ["Total Evaluations", result.total_evaluations],
            ["Total Time (min)", f"{result.total_time_seconds / 60:.1f}"],
            ["Final Generation", self.generation],
            ["", ""],
            ["Best Avg F1", f"{result.best_metrics.get('f1', 0):.4f}"],
            ["Best Model", result.best_metrics.get('best_model', '')],
            ["Best Single F1", f"{result.best_metrics.get('best_f1', 0):.4f}"],
            ["", ""],
            ["Best Features:", ""],
        ]
        for i, feature in enumerate(result.best_features, 1):
            info_data.append([f"  {i}.", feature])

        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)

        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 40

        # Save workbook
        wb.save(excel_path)

        if self.config.verbose:
            print(f"\nExcel saved: {excel_path}")

    def _write_model_sheet(self, wb, sheet_name, records, model_attr, header_font, header_fill, border):
        """Write detailed sheet for a specific model."""
        ws = wb.create_sheet(sheet_name)

        headers = [
            "ID", "Gen", "F1", "Prec", "Rec", "AUC", "Acc",
            "T50_Prec", "T50_Rec", "T50_WR", "T50_PnL", "T50_Cnt",
            "T70_Prec", "T70_Rec", "T70_WR", "T70_PnL", "T70_Cnt",
            "T90_Prec", "T90_Rec", "T90_WR", "T90_PnL", "T90_Cnt",
            "Features"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, record in enumerate(records, 2):
            model = getattr(record, model_attr)
            row_data = [
                record.eval_id, record.generation,
                model.f1, model.precision, model.recall, model.auc, model.accuracy,
                model.thresh_050_prec, model.thresh_050_rec, model.thresh_050_wr, model.thresh_050_pnl, model.thresh_050_cnt,
                model.thresh_070_prec, model.thresh_070_rec, model.thresh_070_wr, model.thresh_070_pnl, model.thresh_070_cnt,
                model.thresh_090_prec, model.thresh_090_rec, model.thresh_090_wr, model.thresh_090_pnl, model.thresh_090_cnt,
                ", ".join(record.features),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '0.0000' if col_idx <= 7 else '0.00'

    def _write_top_summary_sheet(self, ws, records, headers, header_font, header_fill, border):
        """Write top-N summary sheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, record in enumerate(records, 2):
            row_data = [
                record.eval_id, record.generation,
                record.avg_f1, record.avg_precision, record.avg_recall, record.avg_auc,
                record.best_model, record.best_f1,
                record.lgb.f1, record.xgb.f1, record.cb.f1, record.rf.f1, record.lr.f1,
                ", ".join(record.features),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '0.0000'

    def _print_header(self):
        """Print header information."""
        print("\n" + "=" * 80)
        print("GENETIC ALGORITHM FEATURE SELECTION (5 MODELS)")
        print("=" * 80)
        print(f"Population: {self.config.population_size}")
        print(f"Generations: {self.config.n_generations} (early stop: {self.config.early_stopping_rounds})")
        print(f"Features: {self.config.n_features_to_select} from {self.n_features}")
        print(f"Models: LightGBM, XGBoost, CatBoost, RandomForest, LogisticRegression")
        print(f"Fitness: Average F1 across all models")

        # Estimate time
        est_evals = self.config.population_size * self.config.n_generations
        est_time_min = est_evals * 1.5 / 60  # ~1.5s per eval with 5 models
        print(f"Estimated: ~{est_evals} evaluations, ~{est_time_min:.0f} min")
        print("=" * 80)
        print(f"\n{'Gen':>4} | {'%':>5} | {'Best':>7} | {'Avg':>7} | {'Cache':>6} | {'NoImpr':>6} | Best Features")
        print("-" * 80)
        self._start_time = time.time()

    def _print_generation_stats(self):
        """Print current generation statistics with progress."""
        stats = self.history[-1]
        features_str = ", ".join(stats.best_features[:3]) + "..."

        # Progress percentage
        progress = (self.generation + 1) / self.config.n_generations * 100

        # Cache hit rate
        cache_hits = len(self._evaluation_cache)

        # ETA
        elapsed = time.time() - self._start_time
        if self.generation > 0:
            eta_sec = elapsed / (self.generation + 1) * (self.config.n_generations - self.generation - 1)
            eta_str = f"{eta_sec/60:.1f}m" if eta_sec > 60 else f"{eta_sec:.0f}s"
        else:
            eta_str = "..."

        print(
            f"{stats.generation:>4} | "
            f"{progress:>4.1f}% | "
            f"{stats.best_fitness:>7.4f} | "
            f"{stats.avg_fitness:>7.4f} | "
            f"{cache_hits:>6} | "
            f"{self.generations_without_improvement:>6} | "
            f"{features_str} [ETA: {eta_str}]"
        )

    def _print_final_results(self, result: GAResult):
        """Print final results."""
        print("\n" + "=" * 70)
        print("FINAL RESULTS")
        print("=" * 70)
        print(f"Total evaluations: {result.total_evaluations}")
        print(f"Total time: {result.total_time_seconds:.1f}s ({result.total_time_seconds/60:.1f} min)")
        print(f"Final generation: {self.generation}")

        print(f"\nBest {self.config.fitness_metric.upper()}: {result.best_metrics[self.config.fitness_metric]:.4f}")

        print("\nAll metrics:")
        for metric, value in result.best_metrics.items():
            if isinstance(value, float):
                print(f"  {metric}: {value:.4f}")
            else:
                print(f"  {metric}: {value}")

        print(f"\nBest features ({len(result.best_features)}):")
        for i, feature in enumerate(result.best_features, 1):
            print(f"  {i:>2}. {feature}")

        print(f"\nResults saved to: {self.config.output_dir}")
        print("=" * 70)
