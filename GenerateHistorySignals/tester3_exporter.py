# -*- coding: utf-8 -*-
"""
Tester3 Exporter - Export trades in C++ tester_3 format.

Outputs XLSX with exact same structure as C++ tester_3 trades.csv.
Used for SMAEMA strategy comparison.
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import List, Optional

OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    _test_wb = openpyxl.Workbook()
    del _test_wb
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

from models import Trade


# Column definitions matching C++ tester_3 trades.csv
TESTER3_COLUMNS = [
    "signal_id",
    "direction",
    "entry_time",
    "exit_time",
    "entry_price",
    "exit_price",
    "quantity_lots",
    "size_usd",
    "fee_usd",
    "pnl",
    "cum_profit",
    "reason",
    "sl",
    "tp",
]


class Tester3Exporter:
    """Export trades in C++ tester_3 format."""

    def __init__(self, output_path: str):
        """Initialize exporter.

        Args:
            output_path: Path to output XLSX file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def export(
        self,
        trades: List[Trade],
        order_size_usd: float = 100.0,
    ) -> str:
        """Export trades to XLSX in tester_3 format.

        Args:
            trades: List of Trade objects
            order_size_usd: Order size in USD (for quantity calculation)

        Returns:
            Path to saved XLSX file
        """
        ws = self.wb.create_sheet("Trades")

        # Header row with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(TESTER3_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Filter only traded signals (not skipped)
        traded = [t for t in trades if t.trade_status == "traded"]

        # Sort by entry date
        traded_sorted = sorted(traded, key=lambda t: t.signal.date)

        # Track cumulative profit
        cum_profit = 0.0

        # Data rows
        for row_idx, trade in enumerate(traded_sorted, 2):
            row_data = self._build_row(trade, order_size_usd, cum_profit)
            cum_profit = row_data["cum_profit"]

            for col_idx, col_name in enumerate(TESTER3_COLUMNS, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format numbers
                if col_name in ("entry_price", "exit_price", "sl", "tp"):
                    cell.number_format = "#,##0.0"
                elif col_name in ("quantity_lots",):
                    cell.number_format = "0.000"
                elif col_name in ("size_usd", "fee_usd", "pnl", "cum_profit"):
                    cell.number_format = "#,##0.000000"

                # Color reason column
                if col_name == "reason":
                    if value == "TP":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "SL":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "TIMEOUT":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                # Color pnl column
                if col_name == "pnl":
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color="006400")
                        elif value < 0:
                            cell.font = Font(color="8B0000")

        # Auto-fit columns
        for col_idx in range(1, len(TESTER3_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header
        ws.freeze_panes = "A2"

        # Save
        os.makedirs(os.path.dirname(self.output_path) or ".", exist_ok=True)
        self.wb.save(self.output_path)
        return self.output_path

    def _build_row(
        self,
        trade: Trade,
        order_size_usd: float,
        prev_cum_profit: float,
    ) -> dict:
        """Build row data for a single trade.

        Args:
            trade: Trade object
            order_size_usd: Order size in USD
            prev_cum_profit: Cumulative profit before this trade

        Returns:
            Dict with column values
        """
        signal = trade.signal

        # Calculate quantity in lots (coins)
        entry_price = signal.entry
        quantity_lots = order_size_usd / entry_price if entry_price > 0 else 0

        # Calculate actual size based on entry price
        size_usd = quantity_lots * entry_price

        # Calculate fee in USD
        # fee_pct is total fee (entry + exit) as percentage
        fee_usd = size_usd * (trade.fee_pct / 100)

        # Calculate PnL in USD
        pnl_usd = size_usd * (trade.net_pnl_pct / 100)

        # Cumulative profit
        cum_profit = prev_cum_profit + pnl_usd

        # Map result to C++ reason format
        reason_map = {
            "WIN": "TP",
            "LOSS": "SL",
            "TIMEOUT": "TIMEOUT",
        }
        reason = reason_map.get(trade.result, trade.result)

        # Format dates as DD.MM.YYYY HH:MM:SS.000
        entry_time = signal.date.strftime("%d.%m.%Y %H:%M:%S.000")
        exit_time = trade.exit_date.strftime("%d.%m.%Y %H:%M:%S.000")

        # For signal_id, use crossover_idx from metadata if available
        signal_id = signal.metadata.get("crossover_idx", 0)

        return {
            "signal_id": signal_id,
            "direction": signal.direction,
            "entry_time": entry_time,
            "exit_time": exit_time,
            "entry_price": entry_price,
            "exit_price": trade.exit_price,
            "quantity_lots": quantity_lots,
            "size_usd": size_usd,
            "fee_usd": fee_usd,
            "pnl": pnl_usd,
            "cum_profit": cum_profit,
            "reason": reason,
            "sl": signal.stop_loss,
            "tp": signal.take_profit,
        }


def export_tester3_format(
    trades: List[Trade],
    output_path: str,
    order_size_usd: float = 100.0,
) -> str:
    """Convenience function to export trades in tester_3 format.

    Args:
        trades: List of Trade objects
        output_path: Path to output XLSX file
        order_size_usd: Order size in USD

    Returns:
        Path to saved XLSX file
    """
    exporter = Tester3Exporter(output_path)
    return exporter.export(trades, order_size_usd)
