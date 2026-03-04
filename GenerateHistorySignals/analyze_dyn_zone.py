# -*- coding: utf-8 -*-
"""
Analyze WIN-after-WIN pattern in DYN zones.

Проверяем: работает ли паттерн "после WIN 68% WIN" внутри DYN зоны?
"""

import json
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

# Пути
CACHE_DIR = Path("G:/BinanceFriend/GenerateHistorySignals/cache/binance")
OUTPUT_DIR = Path("G:/BinanceFriend/outputNEWARCH/obucheniyeML24_26gg_68_monet_primerno_outputNEWARCH")

# COIN REGIME MATRIX (копия из strategy_runner.py)
COIN_REGIME_MATRIX = {
    'STRONG_BULL': {
        'ls_fade': 'OFF', 'momentum': 'FULL', 'reversal': 'OFF',
        'mean_reversion': 'OFF', 'momentum_ls': 'DYN',
    },
    'BULL': {
        'ls_fade': 'DYN', 'momentum': 'DYN', 'reversal': 'DYN',
        'mean_reversion': 'FULL', 'momentum_ls': 'OFF',
    },
    'SIDEWAYS': {
        'ls_fade': 'FULL', 'momentum': 'OFF', 'reversal': 'DYN',
        'mean_reversion': 'FULL', 'momentum_ls': 'DYN',
    },
    'BEAR': {
        'ls_fade': 'FULL', 'momentum': 'FULL', 'reversal': 'OFF',
        'mean_reversion': 'FULL', 'momentum_ls': 'FULL',
    },
    'STRONG_BEAR': {
        'ls_fade': 'FULL', 'momentum': 'FULL', 'reversal': 'OFF',
        'mean_reversion': 'DYN', 'momentum_ls': 'FULL',
    },
}


def load_klines(symbol: str) -> Dict[str, float]:
    """Загрузить klines: date_str -> close_price"""
    klines_path = CACHE_DIR / symbol / "klines.json"
    if not klines_path.exists():
        return {}

    closes = {}
    with open(klines_path, 'r') as f:
        data = json.load(f)

    for k in data:
        if isinstance(k, list):
            ts = k[0]
            close = float(k[4])
        else:
            ts = k.get('timestamp', k.get('t', 0))
            close = float(k.get('close', k.get('c', 0)))

        if ts > 0:
            dt = datetime.utcfromtimestamp(ts / 1000)
            date_str = dt.strftime('%Y-%m-%d')
            closes[date_str] = close

    return closes


def calculate_regime(closes: Dict[str, float], target_date: str, lookback: int = 14) -> str:
    """Рассчитать режим монеты."""
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
    except:
        return 'UNKNOWN'

    current_close = None
    past_close = None

    # Use PREVIOUS day close (no look-ahead bias)
    for offset in range(1, 4):
        date_str = (target_dt - timedelta(days=offset)).strftime('%Y-%m-%d')
        if date_str in closes:
            current_close = closes[date_str]
            break

    past_dt = target_dt - timedelta(days=lookback)
    for offset in range(1, 4):  # Start from 1 to avoid look-ahead bias
        date_str = (past_dt - timedelta(days=offset)).strftime('%Y-%m-%d')
        if date_str in closes:
            past_close = closes[date_str]
            break

    if current_close is None or past_close is None or past_close == 0:
        return 'UNKNOWN'

    change_pct = (current_close - past_close) / past_close * 100

    if change_pct > 20:
        return 'STRONG_BULL'
    elif change_pct > 5:
        return 'BULL'
    elif change_pct > -5:
        return 'SIDEWAYS'
    elif change_pct > -20:
        return 'BEAR'
    else:
        return 'STRONG_BEAR'


def load_trades_from_xlsx(xlsx_dir: Path) -> List[Dict]:
    """Загрузить все трейды из xlsx файлов."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed")
        return []

    all_trades = []
    xlsx_files = list(xlsx_dir.glob("backtest_*.xlsx"))
    print(f"Found {len(xlsx_files)} xlsx files")

    for xlsx_path in xlsx_files:
        filename = xlsx_path.name
        strategy = None
        for s in ['ls_fade', 'momentum_ls', 'momentum', 'reversal', 'mean_reversion']:
            if s in filename:
                strategy = s
                break

        if not strategy:
            continue

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb['Trades']
            headers = [cell.value for cell in ws[1]]

            date_col = symbol_col = result_col = pnl_col = None
            for i, h in enumerate(headers):
                if h == 'Signal Date': date_col = i
                elif h == 'Symbol': symbol_col = i
                elif h == 'Result': result_col = i
                elif h == 'Net PnL %': pnl_col = i

            if None in [date_col, symbol_col, result_col, pnl_col]:
                wb.close()
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[date_col] is None:
                    continue

                date_val = row[date_col]
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)[:10]

                symbol = row[symbol_col]
                result = row[result_col]
                pnl = row[pnl_col]

                if isinstance(pnl, str):
                    pnl = float(pnl.replace('%', '')) if pnl else 0
                elif pnl is None:
                    pnl = 0
                elif abs(pnl) < 1:
                    pnl = pnl * 100

                all_trades.append({
                    'date': date_str,
                    'symbol': symbol,
                    'strategy': strategy,
                    'result': result,
                    'pnl': float(pnl) if pnl else 0,
                })

            wb.close()
        except Exception as e:
            print(f"Error reading {xlsx_path.name}: {e}")
            continue

    return all_trades


def analyze_dyn_zone():
    """Главная функция анализа."""
    print("=" * 80)
    print("ANALYZE WIN-AFTER-WIN IN DYN ZONE")
    print("=" * 80)
    print()

    # 1. Загрузить трейды
    print("[1/3] Loading trades...")
    trades = load_trades_from_xlsx(OUTPUT_DIR)
    print(f"  Loaded {len(trades)} trades")

    if not trades:
        print("ERROR: No trades found!")
        return

    # 2. Загрузить данные монет
    print("[2/3] Loading coin price data...")
    coin_closes_cache = {}
    symbols_needed = set(t['symbol'] for t in trades)

    for symbol in symbols_needed:
        coin_closes_cache[symbol] = load_klines(symbol)

    loaded = sum(1 for v in coin_closes_cache.values() if v)
    print(f"  Loaded {loaded}/{len(symbols_needed)} symbols")

    # 3. Определить зону для каждого трейда и проанализировать
    print("[3/3] Analyzing zones...")

    # Группируем по (symbol, strategy) и сортируем по дате
    grouped = defaultdict(list)
    for t in trades:
        key = (t['symbol'], t['strategy'])
        grouped[key].append(t)

    # Сортируем каждую группу по дате
    for key in grouped:
        grouped[key].sort(key=lambda x: x['date'])

    # Статистика
    stats = {
        'ALL': {'after_win_win': 0, 'after_win_loss': 0, 'after_loss_win': 0, 'after_loss_loss': 0},
        'FULL': {'after_win_win': 0, 'after_win_loss': 0, 'after_loss_win': 0, 'after_loss_loss': 0},
        'DYN': {'after_win_win': 0, 'after_win_loss': 0, 'after_loss_win': 0, 'after_loss_loss': 0},
        'OFF': {'after_win_win': 0, 'after_win_loss': 0, 'after_loss_win': 0, 'after_loss_loss': 0},
    }

    # Детальная статистика по стратегиям в DYN
    dyn_by_strategy = defaultdict(lambda: {'after_win_win': 0, 'after_win_loss': 0, 'after_loss_win': 0, 'after_loss_loss': 0})

    processed = 0
    skipped = 0

    for (symbol, strategy), trade_list in grouped.items():
        coin_closes = coin_closes_cache.get(symbol, {})
        if not coin_closes:
            skipped += len(trade_list)
            continue

        prev_result = None
        prev_zone = None

        for trade in trade_list:
            # Определить режим и зону
            regime = calculate_regime(coin_closes, trade['date'], lookback=14)

            if regime == 'UNKNOWN' or regime not in COIN_REGIME_MATRIX:
                skipped += 1
                continue

            zone = COIN_REGIME_MATRIX[regime].get(strategy, 'FULL')
            result = trade['result']

            # Пропускаем TIMEOUT - нас интересуют только WIN/LOSS
            if result not in ['WIN', 'LOSS']:
                processed += 1
                continue

            # Анализируем паттерн (если есть предыдущий результат в той же зоне)
            if prev_result is not None and prev_zone == zone:
                if prev_result == 'WIN':
                    if result == 'WIN':
                        stats['ALL']['after_win_win'] += 1
                        stats[zone]['after_win_win'] += 1
                        if zone == 'DYN':
                            dyn_by_strategy[strategy]['after_win_win'] += 1
                    else:
                        stats['ALL']['after_win_loss'] += 1
                        stats[zone]['after_win_loss'] += 1
                        if zone == 'DYN':
                            dyn_by_strategy[strategy]['after_win_loss'] += 1
                else:  # prev_result == 'LOSS'
                    if result == 'WIN':
                        stats['ALL']['after_loss_win'] += 1
                        stats[zone]['after_loss_win'] += 1
                        if zone == 'DYN':
                            dyn_by_strategy[strategy]['after_loss_win'] += 1
                    else:
                        stats['ALL']['after_loss_loss'] += 1
                        stats[zone]['after_loss_loss'] += 1
                        if zone == 'DYN':
                            dyn_by_strategy[strategy]['after_loss_loss'] += 1

            prev_result = result
            prev_zone = zone
            processed += 1

    print(f"  Processed: {processed}, Skipped: {skipped}")
    print()

    # Вывод результатов
    print("=" * 80)
    print("RESULTS: WIN-AFTER-WIN ANALYSIS")
    print("=" * 80)
    print()

    def print_stats(name: str, s: dict):
        total_after_win = s['after_win_win'] + s['after_win_loss']
        total_after_loss = s['after_loss_win'] + s['after_loss_loss']

        if total_after_win > 0:
            pct_win_after_win = s['after_win_win'] / total_after_win * 100
        else:
            pct_win_after_win = 0

        if total_after_loss > 0:
            pct_win_after_loss = s['after_loss_win'] / total_after_loss * 100
        else:
            pct_win_after_loss = 0

        print(f"{name}:")
        print(f"  After WIN:  {s['after_win_win']} WIN / {s['after_win_loss']} LOSS = {pct_win_after_win:.1f}% WIN rate")
        print(f"  After LOSS: {s['after_loss_win']} WIN / {s['after_loss_loss']} LOSS = {pct_win_after_loss:.1f}% WIN rate")
        print(f"  Sample: {total_after_win + total_after_loss} transitions")
        print()

        return pct_win_after_win, pct_win_after_loss, total_after_win + total_after_loss

    print_stats("ALL ZONES", stats['ALL'])
    print_stats("FULL ZONE", stats['FULL'])
    pct_dyn_win, pct_dyn_loss, sample_dyn = print_stats("DYN ZONE", stats['DYN'])

    print("=" * 80)
    print("DYN ZONE BY STRATEGY")
    print("=" * 80)
    print()

    for strategy in ['ls_fade', 'momentum', 'reversal', 'mean_reversion', 'momentum_ls']:
        if strategy in dyn_by_strategy:
            s = dyn_by_strategy[strategy]
            total_win = s['after_win_win'] + s['after_win_loss']
            total_loss = s['after_loss_win'] + s['after_loss_loss']

            if total_win > 0:
                pct = s['after_win_win'] / total_win * 100
                print(f"{strategy:16} After WIN:  {pct:5.1f}% WIN (n={total_win})")
            if total_loss > 0:
                pct = s['after_loss_win'] / total_loss * 100
                print(f"{strategy:16} After LOSS: {pct:5.1f}% WIN (n={total_loss})")
            print()

    print("=" * 80)
    print("CONCLUSION")
    print("=" * 80)
    print()

    if sample_dyn >= 100:
        if pct_dyn_win >= 50:
            print(f"YES! Pattern works in DYN zone: {pct_dyn_win:.1f}% WIN after WIN")
            print()
            print("RECOMMENDATION: Use dynamic sizing in DYN zone")
            print("  - Start with $1")
            print("  - After WIN -> $100")
            print("  - After LOSS -> back to $1")
        else:
            print(f"NO. Pattern does NOT work in DYN zone: only {pct_dyn_win:.1f}% WIN after WIN")
            print()
            print("RECOMMENDATION: Keep DYN zone as fixed $1 or convert to OFF")
    else:
        print(f"NOT ENOUGH DATA: only {sample_dyn} transitions in DYN zone")
        print("Cannot make reliable conclusion")


if __name__ == "__main__":
    analyze_dyn_zone()
