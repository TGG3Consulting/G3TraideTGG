# -*- coding: utf-8 -*-
"""
XLSX Exporter - Export all backtest data to Excel.

Principles:
- ALL field names are CONSISTENT (single naming system)
- ALL API data included (even unused)
- Calculated fields added
"""
from __future__ import annotations

import os
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional

OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    # Test that openpyxl actually works (Python 3.14 compatibility issue)
    _test_wb = openpyxl.Workbook()
    del _test_wb
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

# Import at runtime - these are needed for actual operations
from strategies import StrategyConfig, DailyCandle
from data_downloader import SymbolHistoryData
from models import Trade, BacktestResult


# CONSISTENT FIELD NAMES (used everywhere!)
FIELD_NAMES = {
    # Candle - Entry Day (all fields from Binance klines)
    "candle_timestamp": "Candle Time",
    "candle_open": "Open",
    "candle_high": "High",
    "candle_low": "Low",
    "candle_close": "Close",
    "candle_volume": "Volume",
    "candle_quote_volume": "Volume USD",
    "candle_trades_count": "Trades Count",
    "candle_taker_buy_volume": "Taker Buy Vol",
    "candle_taker_buy_quote_volume": "Taker Buy USD",

    # Candle - Previous Day (for ML - no look-ahead bias)
    "prev_high": "Prev High",
    "prev_low": "Prev Low",
    "prev_close": "Prev Close",
    "prev_volume": "Prev Volume",
    "prev_quote_volume": "Prev Volume USD",
    "prev_trades_count": "Prev Trades Count",
    "prev_taker_buy_volume": "Prev Taker Buy Vol",
    "prev_taker_buy_quote_volume": "Prev Taker Buy USD",

    # L/S Ratio
    "ls_long_pct": "Long %",
    "ls_short_pct": "Short %",
    "ls_ratio": "L/S Ratio",

    # Open Interest
    "oi_contracts": "OI Contracts",
    "oi_value_usd": "OI USD",

    # Funding
    "funding_rate": "Funding Rate",

    # Signal
    "signal_date": "Signal Date",
    "signal_symbol": "Symbol",
    "signal_direction": "Direction",
    "signal_entry": "Entry Price",
    "signal_sl": "Stop Loss",
    "signal_tp": "Take Profit",
    "signal_reason": "Reason",

    # Trade Result
    "trade_exit_date": "Exit Date",
    "trade_exit_price": "Exit Price",
    "trade_gross_pnl_pct": "Gross PnL %",
    "trade_fee_pct": "Fee %",
    "trade_funding_fee_pct": "Funding Fee %",
    "trade_funding_periods": "Funding Periods",
    "trade_slippage_pct": "Slippage %",
    "trade_net_pnl_pct": "Net PnL %",
    "trade_result": "Result",
    "trade_hold_days": "Hold Days",

    # Calculated
    "calc_sl_pct": "SL %",
    "calc_tp_pct": "TP %",
    "calc_rr_ratio": "R:R Ratio",
    "calc_pnl_usd": "PnL USD",
    "calc_volume_ratio": "Vol Ratio",
    "current_dd_pct": "Current DD %",
    "cumulative_pnl_pct": "Cumulative PnL %",
    "adx": "ADX",
    "order_size": "Order Size $",

    # Regime and Volatility
    "coin_regime": "Coin Regime",
    "coin_volatility": "Coin Vol %",
    "atr_pct": "ATR %",

    # Chain/Dedup
    "signal_id": "Signal ID",
    "chain_id": "Chain ID",
    "chain_seq": "Chain Seq",
    "chain_total": "Chain Total",
    "chain_gap_days": "Gap Days",
    "is_chain_first": "Chain First",
    "is_chain_last": "Chain Last",
}

# Column order for Trades sheet (ALL data as per PLAN_IMPROVEMENTS.md)
TRADES_COLUMNS = [
    # Signal identification
    "signal_id",
    "chain_id",
    "chain_seq",
    "chain_total",
    # Signal details
    "signal_date",
    "signal_symbol",
    "signal_direction",
    "signal_entry",
    "signal_sl",
    "signal_tp",
    "calc_sl_pct",
    "calc_tp_pct",
    "calc_rr_ratio",
    # Trade results
    "trade_exit_date",
    "trade_exit_price",
    "trade_gross_pnl_pct",
    "trade_fee_pct",
    "trade_funding_fee_pct",
    "trade_funding_periods",
    "trade_slippage_pct",
    "trade_net_pnl_pct",
    "trade_result",
    "trade_hold_days",
    "order_size",
    "calc_pnl_usd",
    "cumulative_pnl_pct",
    "current_dd_pct",
    # Candle data - Entry Day (all from Binance API)
    "candle_open",
    "candle_high",
    "candle_low",
    "candle_close",
    "candle_volume",
    "candle_quote_volume",
    "candle_trades_count",
    "candle_taker_buy_volume",
    "candle_taker_buy_quote_volume",
    "calc_volume_ratio",
    # Candle data - Previous Day (for ML - no look-ahead bias)
    "prev_high",
    "prev_low",
    "prev_close",
    "prev_volume",
    "prev_quote_volume",
    "prev_trades_count",
    "prev_taker_buy_volume",
    "prev_taker_buy_quote_volume",
    # Market data
    "ls_long_pct",
    "ls_short_pct",
    "oi_contracts",
    "oi_value_usd",
    "funding_rate",
    "adx",
    # Regime and Volatility
    "coin_regime",
    "coin_volatility",
    "atr_pct",
    # Chain metadata
    "chain_gap_days",
    "is_chain_first",
    "is_chain_last",
    "signal_reason",
]


class XLSXExporter:
    """Export backtest results to XLSX with all data."""

    def __init__(self, output_path: str):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    def export_backtest(
        self,
        trades: List[Trade],
        history: Dict[str, SymbolHistoryData],
        config: StrategyConfig,
        result: BacktestResult,
        order_size_usd: float = 100.0,
        strategy_name: str = "unknown",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        market_regime: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Export ALL data to XLSX.

        Args:
            trades: List of completed trades
            history: Historical data by symbol
            config: Strategy configuration
            result: Backtest result summary
            order_size_usd: Order size in USD
            strategy_name: Name of strategy
            start_date: Backtest start date
            end_date: Backtest end date
            market_regime: Market regime detection result

        Returns:
            Path to saved XLSX file
        """
        self._write_trades_sheet(trades, history, order_size_usd)
        self._write_summary_sheet(result, order_size_usd, market_regime)
        self._write_config_sheet(config, strategy_name, start_date, end_date, order_size_usd, market_regime)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(self.output_path) or ".", exist_ok=True)
        self.wb.save(self.output_path)
        return self.output_path

    def _write_trades_sheet(
        self,
        trades: List[Trade],
        history: Dict[str, SymbolHistoryData],
        order_size_usd: float,
    ) -> None:
        """Write Trades sheet with all trade data."""
        ws = self.wb.create_sheet("Trades")

        # Build candle cache for lookups
        candles_cache = self._build_candles_cache(history)

        # Header row
        headers = [FIELD_NAMES.get(col, col) for col in TRADES_COLUMNS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Sort trades by date for proper DD calculation
        sorted_trades = sorted(trades, key=lambda t: t.signal.date)

        # Track equity and drawdown
        cumulative_pnl = 0.0
        peak_equity = 0.0

        # Data rows
        for row_idx, trade in enumerate(sorted_trades, 2):
            row_data = self._build_trade_row(trade, history, candles_cache, order_size_usd)

            # Calculate cumulative PnL and drawdown (only for traded signals)
            if trade.trade_status == "traded":
                cumulative_pnl += trade.net_pnl_pct
                if cumulative_pnl > peak_equity:
                    peak_equity = cumulative_pnl
                current_dd = peak_equity - cumulative_pnl
            else:
                current_dd = peak_equity - cumulative_pnl  # Same DD for skipped

            row_data["cumulative_pnl_pct"] = cumulative_pnl
            row_data["current_dd_pct"] = current_dd

            for col_idx, col_name in enumerate(TRADES_COLUMNS, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format numbers
                if isinstance(value, float):
                    if "pct" in col_name or "ratio" in col_name:
                        cell.number_format = "0.00%"
                        cell.value = value / 100  # Convert to decimal for %
                    elif "usd" in col_name or "price" in col_name or "volume" in col_name:
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "0.0000"

                # Color WIN/LOSS
                if col_name == "trade_result":
                    if value == "WIN":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "LOSS":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "TIMEOUT":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                # Color PnL
                if col_name == "trade_net_pnl_pct":
                    if isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(color="006400")
                    elif isinstance(value, (int, float)) and value < 0:
                        cell.font = Font(color="8B0000")

        # Auto-fit columns
        for col_idx in range(1, len(TRADES_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # Freeze header
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, result: BacktestResult, order_size_usd: float, market_regime: Optional[Dict[str, Any]] = None) -> None:
        """Write Summary sheet with aggregated metrics."""
        ws = self.wb.create_sheet("Summary")

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        ws.cell(row=1, column=1, value="Metric").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="Value").font = header_font_white
        ws.cell(row=1, column=2).fill = header_fill

        # Metrics
        metrics = [
            ("Total Signals", result.total_signals),
            ("Total Trades", result.total_trades),
            ("Skipped (Liquidity)", result.skipped_liquidity),
            ("Skipped (Position)", result.skipped_position),
            ("Position Mode", result.position_mode),
            ("Wins", result.wins),
            ("Losses", result.losses),
            ("Timeouts", result.timeouts),
            ("Win Rate %", f"{result.win_rate:.1f}%"),
            ("Total Net PnL %", f"{result.total_pnl:+.2f}%"),
            ("Avg Net PnL %", f"{result.avg_pnl:+.2f}%"),
            ("Long PnL %", f"{result.long_pnl:+.2f}%"),
            ("Short PnL %", f"{result.short_pnl:+.2f}%"),
            ("Total Fees %", f"{result.total_fees_pct:.2f}%"),
            ("Max Drawdown %", f"{result.max_drawdown:.2f}%"),
            ("Calmar Ratio", f"{result.calmar_ratio:.2f}"),
            ("Avg Hold Win (days)", f"{result.avg_hold_win:.1f}"),
            ("Avg Hold Loss (days)", f"{result.avg_hold_loss:.1f}"),
            ("Avg Hold Timeout (days)", f"{result.avg_hold_timeout:.1f}"),
            ("Order Size USD", f"${order_size_usd:.0f}"),
            ("Total PnL USD", f"${order_size_usd * result.total_pnl / 100:.2f}"),
        ]

        # Add market regime info
        if market_regime:
            metrics.append(("", ""))  # Empty row separator
            metrics.append(("Market Regime", market_regime.get('regime', 'UNKNOWN')))
            metrics.append(("Reference Symbol", market_regime.get('ref_symbol', 'N/A')))
            metrics.append(("Price Change %", f"{market_regime.get('change_pct', 0):+.1f}%"))
            metrics.append(("Avg Volatility %", f"{market_regime.get('volatility', 0):.1f}%"))

        for row_idx, (name, value) in enumerate(metrics, 2):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _write_config_sheet(
        self,
        config: StrategyConfig,
        strategy_name: str,
        start_date: Optional[datetime],
        end_date: Optional[datetime],
        order_size_usd: float,
        market_regime: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Write Config sheet with run parameters."""
        ws = self.wb.create_sheet("Config")

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        ws.cell(row=1, column=1, value="Parameter").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="Value").font = header_font_white
        ws.cell(row=1, column=2).fill = header_fill

        params = [
            ("Strategy", strategy_name),
            ("Start Date", start_date.strftime("%Y-%m-%d") if start_date else "N/A"),
            ("End Date", end_date.strftime("%Y-%m-%d") if end_date else "N/A"),
            ("Order Size USD", order_size_usd),
            ("Stop Loss %", config.sl_pct),
            ("Take Profit %", config.tp_pct),
            ("Max Hold Days", config.max_hold_days),
            ("Lookback Days", config.lookback),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]

        # Add market regime info
        if market_regime:
            params.append(("", ""))
            params.append(("Market Regime", market_regime.get('regime', 'UNKNOWN')))
            params.append(("Reference Symbol", market_regime.get('ref_symbol', 'N/A')))
            params.append(("Price Change %", f"{market_regime.get('change_pct', 0):+.1f}%"))
            params.append(("Avg Volatility %", f"{market_regime.get('volatility', 0):.1f}%"))

        # Add strategy-specific params
        for key, value in config.params.items():
            params.append((f"param_{key}", value))

        for row_idx, (name, value) in enumerate(params, 2):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _build_candles_cache(
        self,
        history: Dict[str, SymbolHistoryData],
    ) -> Dict[str, Dict[str, DailyCandle]]:
        """Build candle lookup cache by symbol and date."""
        cache = {}
        for symbol, data in history.items():
            candles = self._aggregate_to_daily(data.klines)
            cache[symbol] = {c.date.strftime("%Y-%m-%d"): c for c in candles}
        return cache

    @staticmethod
    def _aggregate_to_daily(klines: List[Dict]) -> List[DailyCandle]:
        """Aggregate 1-minute klines to daily candles with ALL available data."""
        from datetime import timezone as tz
        daily = {}

        for k in klines:
            ts = k.get("timestamp", 0)
            dt = datetime.fromtimestamp(ts / 1000, tz=tz.utc)
            date_key = dt.strftime("%Y-%m-%d")

            qv = float(k.get("quote_volume", 0)) or float(k["close"]) * float(k["volume"])
            trades = int(k.get("trades_count", 0))
            taker_buy_vol = float(k.get("taker_buy_volume", 0))
            taker_buy_quote = float(k.get("taker_buy_quote_volume", 0))

            if date_key not in daily:
                daily[date_key] = {
                    "date": dt.replace(hour=0, minute=0, second=0),
                    "open": float(k["open"]),
                    "high": float(k["high"]),
                    "low": float(k["low"]),
                    "close": float(k["close"]),
                    "volume": float(k["volume"]),
                    "quote_volume": qv,
                    "trades_count": trades,
                    "taker_buy_volume": taker_buy_vol,
                    "taker_buy_quote_volume": taker_buy_quote,
                }
            else:
                daily[date_key]["high"] = max(daily[date_key]["high"], float(k["high"]))
                daily[date_key]["low"] = min(daily[date_key]["low"], float(k["low"]))
                daily[date_key]["close"] = float(k["close"])
                daily[date_key]["volume"] += float(k["volume"])
                daily[date_key]["quote_volume"] += qv
                daily[date_key]["trades_count"] += trades
                daily[date_key]["taker_buy_volume"] += taker_buy_vol
                daily[date_key]["taker_buy_quote_volume"] += taker_buy_quote

        candles = []
        for date_key in sorted(daily.keys()):
            d = daily[date_key]
            candles.append(DailyCandle(
                date=d["date"],
                open=d["open"],
                high=d["high"],
                low=d["low"],
                close=d["close"],
                volume=d["volume"],
                quote_volume=d["quote_volume"],
                trades_count=d["trades_count"],
                taker_buy_volume=d["taker_buy_volume"],
                taker_buy_quote_volume=d["taker_buy_quote_volume"],
            ))

        return candles

    def _build_trade_row(
        self,
        trade: Trade,
        history: Dict[str, SymbolHistoryData],
        candles_cache: Dict[str, Dict[str, DailyCandle]],
        order_size_usd: float,
    ) -> Dict[str, Any]:
        """Build a row dict for a single trade."""
        signal = trade.signal
        symbol = signal.symbol
        signal_date_str = signal.date.strftime("%Y-%m-%d")

        row = {
            # Signal
            "signal_date": signal.date.strftime("%Y-%m-%d"),
            "signal_symbol": symbol,
            "signal_direction": signal.direction,
            "signal_entry": signal.entry,
            "signal_sl": signal.stop_loss,
            "signal_tp": signal.take_profit,
            "signal_reason": signal.reason,

            # Chain/Dedup
            "signal_id": signal.signal_id,
            "chain_id": signal.chain_id,
            "chain_seq": signal.chain_seq,
            "chain_total": signal.chain_total,
            "chain_gap_days": signal.chain_gap_days,
            "is_chain_first": signal.is_chain_first,
            "is_chain_last": signal.is_chain_last,

            # Trade
            "trade_exit_date": trade.exit_date.strftime("%Y-%m-%d"),
            "trade_exit_price": trade.exit_price,
            "trade_gross_pnl_pct": trade.pnl_pct,
            "trade_fee_pct": trade.fee_pct,
            "trade_funding_fee_pct": trade.funding_fee_pct,
            "trade_funding_periods": trade.funding_periods,
            "trade_slippage_pct": trade.slippage_pct,
            "trade_net_pnl_pct": trade.net_pnl_pct,
            "trade_result": trade.result,
            "trade_hold_days": trade.hold_days,

            # Order size (dynamic sizing support)
            "order_size": trade.order_size,

            # Regime and Volatility
            "coin_regime": trade.coin_regime,
            "coin_volatility": trade.coin_volatility,
            "atr_pct": trade.atr_pct,

            # Calculated (use actual order_size for PnL calculation)
            "calc_pnl_usd": trade.order_size * trade.net_pnl_pct / 100,
        }

        # Calculate SL/TP percentages
        if signal.direction == "LONG":
            row["calc_sl_pct"] = (signal.entry - signal.stop_loss) / signal.entry * 100
            row["calc_tp_pct"] = (signal.take_profit - signal.entry) / signal.entry * 100
        else:
            row["calc_sl_pct"] = (signal.stop_loss - signal.entry) / signal.entry * 100
            row["calc_tp_pct"] = (signal.entry - signal.take_profit) / signal.entry * 100

        # R:R Ratio
        if row["calc_sl_pct"] > 0:
            row["calc_rr_ratio"] = row["calc_tp_pct"] / row["calc_sl_pct"]
        else:
            row["calc_rr_ratio"] = 0

        # Candle data - Entry Day (ALL fields from Binance klines)
        if symbol in candles_cache and signal_date_str in candles_cache[symbol]:
            candle = candles_cache[symbol][signal_date_str]
            row["candle_open"] = candle.open
            row["candle_high"] = candle.high
            row["candle_low"] = candle.low
            row["candle_close"] = candle.close
            row["candle_volume"] = candle.volume
            row["candle_quote_volume"] = candle.quote_volume
            row["candle_trades_count"] = candle.trades_count
            row["candle_taker_buy_volume"] = candle.taker_buy_volume
            row["candle_taker_buy_quote_volume"] = candle.taker_buy_quote_volume

            # Volume ratio (order size / daily volume)
            if candle.quote_volume > 0:
                row["calc_volume_ratio"] = trade.order_size / candle.quote_volume * 100
            else:
                row["calc_volume_ratio"] = 0

        # Candle data - Previous Day (for ML - no look-ahead bias)
        # Get previous day's date
        from datetime import timedelta
        prev_date = signal.date - timedelta(days=1)
        prev_date_str = prev_date.strftime("%Y-%m-%d")

        if symbol in candles_cache and prev_date_str in candles_cache[symbol]:
            prev_candle = candles_cache[symbol][prev_date_str]
            row["prev_high"] = prev_candle.high
            row["prev_low"] = prev_candle.low
            row["prev_close"] = prev_candle.close
            row["prev_volume"] = prev_candle.volume
            row["prev_quote_volume"] = prev_candle.quote_volume
            row["prev_trades_count"] = prev_candle.trades_count
            row["prev_taker_buy_volume"] = prev_candle.taker_buy_volume
            row["prev_taker_buy_quote_volume"] = prev_candle.taker_buy_quote_volume

        # L/S data
        if symbol in history:
            ls_data = self._get_ls_for_date(history[symbol].ls_ratio_history, signal.date)
            if ls_data:
                row["ls_long_pct"] = float(ls_data.get("longAccount", 0)) * 100
                row["ls_short_pct"] = float(ls_data.get("shortAccount", 0)) * 100

            # OI data (both contracts and USD value)
            oi_data = self._get_oi_for_date(history[symbol].oi_history, signal.date)
            if oi_data:
                row["oi_contracts"] = float(oi_data.get("sumOpenInterest", 0))
                row["oi_value_usd"] = float(oi_data.get("sumOpenInterestValue", 0))

            # Funding data
            funding_data = self._get_funding_for_date(history[symbol].funding_history, signal.date)
            if funding_data:
                row["funding_rate"] = float(funding_data.get("fundingRate", 0)) * 100

            # ADX from signal metadata
            row["adx"] = signal.metadata.get("adx", 0.0)

        return row

    def _get_ls_for_date(self, ls_history: List[Dict], target_date: datetime) -> Optional[Dict]:
        """Get L/S ratio data for a specific date."""
        target_ts = int(target_date.timestamp() * 1000)
        best = None
        best_diff = float('inf')

        for ls in ls_history:
            ts = ls.get("timestamp", 0)
            diff = abs(ts - target_ts)
            if diff < best_diff and ts <= target_ts:
                best_diff = diff
                best = ls

        return best

    def _get_oi_for_date(self, oi_history: List[Dict], target_date: datetime) -> Optional[Dict]:
        """Get OI data for a specific date."""
        target_ts = int(target_date.timestamp() * 1000)
        best = None
        best_diff = float('inf')

        for oi in oi_history:
            ts = oi.get("timestamp", 0)
            diff = abs(ts - target_ts)
            if diff < best_diff and ts <= target_ts:
                best_diff = diff
                best = oi

        return best

    def _get_funding_for_date(self, funding_history: List[Dict], target_date: datetime) -> Optional[Dict]:
        """Get funding rate data for a specific date."""
        target_ts = int(target_date.timestamp() * 1000)
        best = None
        best_diff = float('inf')

        for f in funding_history:
            ts = f.get("fundingTime", 0)
            diff = abs(ts - target_ts)
            if diff < best_diff and ts <= target_ts:
                best_diff = diff
                best = f

        return best
