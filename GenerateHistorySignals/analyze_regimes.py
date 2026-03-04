# -*- coding: utf-8 -*-
"""
Analyze Market Regimes vs Strategy Performance.

Для каждого трейда рассчитывает:
1. BTC Regime (14d lookback)
2. Coin Regime (14d lookback)
3. BTC-Coin Correlation (14d)

Затем строит статистику: regime × strategy → WR%, PnL%, MaxDD%
"""

import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import math

# Пути
CACHE_DIR = Path("G:/BinanceFriend/GenerateHistorySignals/cache/binance")
OUTPUT_DIR = Path("G:/BinanceFriend/outputNEWARCH/obucheniyeML24_26gg_68_monet_primerno_outputNEWARCH")

# Пороги режимов (14d price change %)
REGIME_THRESHOLDS = {
    'STRONG_BULL': 20,
    'BULL': 5,
    'SIDEWAYS_UP': 0,
    'SIDEWAYS_DOWN': -5,
    'BEAR': -20,
    # < -20 = STRONG_BEAR
}

def load_klines(symbol: str) -> Dict[str, float]:
    """
    Загрузить klines и вернуть dict: date_str -> close_price
    """
    klines_path = CACHE_DIR / symbol / "klines.json"
    if not klines_path.exists():
        return {}

    closes = {}
    with open(klines_path, 'r') as f:
        data = json.load(f)

    for k in data:
        # Формат: [timestamp, open, high, low, close, volume, ...]
        # Или dict с ключами
        if isinstance(k, list):
            ts = k[0]
            close = float(k[4])
        else:
            ts = k.get('timestamp', k.get('t', 0))
            close = float(k.get('close', k.get('c', 0)))

        if ts > 0:
            dt = datetime.utcfromtimestamp(ts / 1000)
            date_str = dt.strftime('%Y-%m-%d')
            closes[date_str] = close

    return closes


def calculate_regime(closes: Dict[str, float], target_date: str, lookback: int = 14) -> Tuple[str, float]:
    """
    Рассчитать режим на основе изменения цены за lookback дней.

    Returns:
        (regime_name, change_pct)
    """
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
    except:
        return ('UNKNOWN', 0.0)

    # Найти цену на target_date и lookback дней назад
    current_close = None
    past_close = None

    # Use PREVIOUS day close (no look-ahead bias)
    for offset in range(1, 4):
        date_str = (target_dt - timedelta(days=offset)).strftime('%Y-%m-%d')
        if date_str in closes:
            current_close = closes[date_str]
            break

    # Ищем прошлую цену
    past_dt = target_dt - timedelta(days=lookback)
    for offset in range(1, 4):  # Start from 1 to avoid look-ahead bias
        date_str = (past_dt - timedelta(days=offset)).strftime('%Y-%m-%d')
        if date_str in closes:
            past_close = closes[date_str]
            break

    if current_close is None or past_close is None or past_close == 0:
        return ('UNKNOWN', 0.0)

    change_pct = (current_close - past_close) / past_close * 100

    # Определить режим
    if change_pct > 20:
        regime = 'STRONG_BULL'
    elif change_pct > 5:
        regime = 'BULL'
    elif change_pct > -5:
        regime = 'SIDEWAYS'
    elif change_pct > -20:
        regime = 'BEAR'
    else:
        regime = 'STRONG_BEAR'

    return (regime, change_pct)


def calculate_correlation(btc_closes: Dict[str, float], coin_closes: Dict[str, float],
                          target_date: str, lookback: int = 14) -> float:
    """
    Рассчитать корреляцию daily returns между BTC и монетой.
    """
    try:
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
    except:
        return 0.0

    btc_returns = []
    coin_returns = []

    for i in range(lookback):
        date1 = (target_dt - timedelta(days=i)).strftime('%Y-%m-%d')
        date2 = (target_dt - timedelta(days=i+1)).strftime('%Y-%m-%d')

        if date1 in btc_closes and date2 in btc_closes and date1 in coin_closes and date2 in coin_closes:
            btc_ret = (btc_closes[date1] - btc_closes[date2]) / btc_closes[date2]
            coin_ret = (coin_closes[date1] - coin_closes[date2]) / coin_closes[date2]
            btc_returns.append(btc_ret)
            coin_returns.append(coin_ret)

    if len(btc_returns) < 5:
        return 0.0

    # Pearson correlation
    n = len(btc_returns)
    mean_btc = sum(btc_returns) / n
    mean_coin = sum(coin_returns) / n

    cov = sum((btc_returns[i] - mean_btc) * (coin_returns[i] - mean_coin) for i in range(n))
    std_btc = math.sqrt(sum((x - mean_btc) ** 2 for x in btc_returns))
    std_coin = math.sqrt(sum((x - mean_coin) ** 2 for x in coin_returns))

    if std_btc == 0 or std_coin == 0:
        return 0.0

    correlation = cov / (std_btc * std_coin)
    return correlation


def load_trades_from_xlsx(xlsx_dir: Path) -> List[Dict]:
    """
    Загрузить все трейды из xlsx файлов.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed")
        return []

    all_trades = []

    # Найти все xlsx файлы
    xlsx_files = list(xlsx_dir.glob("backtest_*.xlsx"))
    print(f"Found {len(xlsx_files)} xlsx files")

    for xlsx_path in xlsx_files:
        # Определить стратегию из имени файла
        filename = xlsx_path.name
        strategy = None
        for s in ['ls_fade', 'momentum_ls', 'momentum', 'reversal', 'mean_reversion']:
            if s in filename:
                strategy = s
                break

        if not strategy:
            continue

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb['Trades']

            # Получить заголовки
            headers = [cell.value for cell in ws[1]]

            # Найти нужные колонки
            date_col = None
            symbol_col = None
            result_col = None
            pnl_col = None

            for i, h in enumerate(headers):
                if h == 'Signal Date':
                    date_col = i
                elif h == 'Symbol':
                    symbol_col = i
                elif h == 'Result':
                    result_col = i
                elif h == 'Net PnL %':
                    pnl_col = i

            if None in [date_col, symbol_col, result_col, pnl_col]:
                wb.close()
                continue

            # Читать трейды
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[date_col] is None:
                    continue

                date_val = row[date_col]
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)[:10]

                symbol = row[symbol_col]
                result = row[result_col]
                pnl = row[pnl_col]

                if isinstance(pnl, str):
                    pnl = float(pnl.replace('%', '')) if pnl else 0
                elif pnl is None:
                    pnl = 0
                else:
                    # Может быть в формате 0.05 (5%)
                    if abs(pnl) < 1:
                        pnl = pnl * 100

                all_trades.append({
                    'date': date_str,
                    'symbol': symbol,
                    'strategy': strategy,
                    'result': result,
                    'pnl': float(pnl) if pnl else 0,
                })

            wb.close()

        except Exception as e:
            print(f"Error reading {xlsx_path.name}: {e}")
            continue

    return all_trades


def analyze_regimes():
    """
    Главная функция анализа.
    """
    print("=" * 80)
    print("MARKET REGIME ANALYSIS")
    print("=" * 80)
    print()

    # 1. Загрузить BTC данные
    print("[1/4] Loading BTC price data...")
    btc_closes = load_klines('BTCUSDT')
    print(f"  BTC: {len(btc_closes)} daily closes")

    if not btc_closes:
        print("ERROR: No BTC data found!")
        return

    # 2. Загрузить трейды
    print("[2/4] Loading trades from xlsx...")
    trades = load_trades_from_xlsx(OUTPUT_DIR)
    print(f"  Loaded {len(trades)} trades")

    if not trades:
        print("ERROR: No trades found!")
        return

    # 3. Загрузить данные монет (по мере необходимости)
    print("[3/4] Loading coin price data...")
    coin_closes_cache = {}
    symbols_needed = set(t['symbol'] for t in trades)

    for symbol in symbols_needed:
        coin_closes_cache[symbol] = load_klines(symbol)

    loaded_symbols = sum(1 for v in coin_closes_cache.values() if v)
    print(f"  Loaded {loaded_symbols}/{len(symbols_needed)} symbols")

    # 4. Рассчитать режимы для каждого трейда
    print("[4/4] Calculating regimes for each trade...")

    # Статистика: (btc_regime, coin_regime, strategy) -> [trades]
    stats = defaultdict(list)
    stats_btc_only = defaultdict(list)
    stats_coin_only = defaultdict(list)

    processed = 0
    skipped = 0

    for trade in trades:
        date_str = trade['date']
        symbol = trade['symbol']
        strategy = trade['strategy']
        pnl = trade['pnl']
        result = trade['result']

        # BTC regime
        btc_regime, btc_change = calculate_regime(btc_closes, date_str, lookback=14)

        # Coin regime
        coin_closes = coin_closes_cache.get(symbol, {})
        if coin_closes:
            coin_regime, coin_change = calculate_regime(coin_closes, date_str, lookback=14)
            correlation = calculate_correlation(btc_closes, coin_closes, date_str, lookback=14)
        else:
            coin_regime, coin_change = 'UNKNOWN', 0.0
            correlation = 0.0

        if btc_regime == 'UNKNOWN':
            skipped += 1
            continue

        # Записать в статистику
        trade_data = {
            'pnl': pnl,
            'result': result,
            'btc_change': btc_change,
            'coin_change': coin_change,
            'correlation': correlation,
        }

        # Полная статистика (BTC + Coin)
        key = (btc_regime, coin_regime, strategy)
        stats[key].append(trade_data)

        # Только BTC regime
        key_btc = (btc_regime, strategy)
        stats_btc_only[key_btc].append(trade_data)

        # Только Coin regime
        if coin_regime != 'UNKNOWN':
            key_coin = (coin_regime, strategy)
            stats_coin_only[key_coin].append(trade_data)

        processed += 1

        if processed % 10000 == 0:
            print(f"  Processed {processed}/{len(trades)}...")

    print(f"  Done: {processed} processed, {skipped} skipped")
    print()

    # 5. Вывести результаты
    print("=" * 80)
    print("RESULTS: BTC REGIME x STRATEGY")
    print("=" * 80)
    print()
    print(f"{'BTC Regime':<12} {'Strategy':<15} {'Trades':>7} {'WinRate':>8} {'AvgPnL':>8} {'TotalPnL':>10} {'MaxDD':>8}")
    print("-" * 80)

    # Сортировка по режиму
    regime_order = ['STRONG_BULL', 'BULL', 'SIDEWAYS', 'BEAR', 'STRONG_BEAR']
    strategy_order = ['ls_fade', 'momentum', 'reversal', 'mean_reversion', 'momentum_ls']

    results_btc = []

    for regime in regime_order:
        for strategy in strategy_order:
            key = (regime, strategy)
            if key not in stats_btc_only:
                continue

            trades_list = stats_btc_only[key]
            n = len(trades_list)
            wins = sum(1 for t in trades_list if t['result'] == 'WIN')
            wr = wins / n * 100 if n > 0 else 0
            avg_pnl = sum(t['pnl'] for t in trades_list) / n if n > 0 else 0
            total_pnl = sum(t['pnl'] for t in trades_list)

            # MaxDD (простой расчет)
            cumsum = 0
            peak = 0
            maxdd = 0
            for t in trades_list:
                cumsum += t['pnl']
                if cumsum > peak:
                    peak = cumsum
                dd = peak - cumsum
                if dd > maxdd:
                    maxdd = dd

            results_btc.append({
                'regime': regime,
                'strategy': strategy,
                'trades': n,
                'wr': wr,
                'avg_pnl': avg_pnl,
                'total_pnl': total_pnl,
                'maxdd': -maxdd,
            })

            print(f"{regime:<12} {strategy:<15} {n:>7} {wr:>7.1f}% {avg_pnl:>+7.2f}% {total_pnl:>+9.1f}% {-maxdd:>7.1f}%")

    print()
    print("=" * 80)
    print("RESULTS: COIN REGIME x STRATEGY")
    print("=" * 80)
    print()
    print(f"{'Coin Regime':<12} {'Strategy':<15} {'Trades':>7} {'WinRate':>8} {'AvgPnL':>8} {'TotalPnL':>10} {'MaxDD':>8}")
    print("-" * 80)

    for regime in regime_order:
        for strategy in strategy_order:
            key = (regime, strategy)
            if key not in stats_coin_only:
                continue

            trades_list = stats_coin_only[key]
            n = len(trades_list)
            wins = sum(1 for t in trades_list if t['result'] == 'WIN')
            wr = wins / n * 100 if n > 0 else 0
            avg_pnl = sum(t['pnl'] for t in trades_list) / n if n > 0 else 0
            total_pnl = sum(t['pnl'] for t in trades_list)

            # MaxDD
            cumsum = 0
            peak = 0
            maxdd = 0
            for t in trades_list:
                cumsum += t['pnl']
                if cumsum > peak:
                    peak = cumsum
                dd = peak - cumsum
                if dd > maxdd:
                    maxdd = dd

            print(f"{regime:<12} {strategy:<15} {n:>7} {wr:>7.1f}% {avg_pnl:>+7.2f}% {total_pnl:>+9.1f}% {-maxdd:>7.1f}%")

    print()
    print("=" * 80)
    print("KEY INSIGHTS")
    print("=" * 80)
    print()

    # Найти лучшие/худшие комбинации
    print("TOP 5 - Best combinations (by WinRate, min 100 trades):")
    sorted_btc = sorted([r for r in results_btc if r['trades'] >= 100],
                        key=lambda x: x['wr'], reverse=True)[:5]
    for r in sorted_btc:
        print(f"  {r['regime']} + {r['strategy']}: {r['wr']:.1f}% WR, {r['total_pnl']:+.1f}% PnL ({r['trades']} trades)")

    print()
    print("BOTTOM 5 - Worst combinations (by WinRate, min 100 trades):")
    sorted_btc_worst = sorted([r for r in results_btc if r['trades'] >= 100],
                              key=lambda x: x['wr'])[:5]
    for r in sorted_btc_worst:
        print(f"  {r['regime']} + {r['strategy']}: {r['wr']:.1f}% WR, {r['total_pnl']:+.1f}% PnL ({r['trades']} trades)")

    print()
    print("=" * 80)
    print("RECOMMENDED MATRIX FOR LIVE TRADING")
    print("=" * 80)
    print()
    print("Based on WinRate and MaxDD analysis:")
    print()
    print(f"{'BTC Regime':<12} | {'ls_fade':^10} | {'momentum':^10} | {'reversal':^10} | {'mean_rev':^10} | {'mom_ls':^10}")
    print("-" * 80)

    # Построить рекомендации
    for regime in regime_order:
        row = f"{regime:<12} |"
        for strategy in strategy_order:
            key = (regime, strategy)
            if key in stats_btc_only:
                trades_list = stats_btc_only[key]
                n = len(trades_list)
                if n >= 50:
                    wins = sum(1 for t in trades_list if t['result'] == 'WIN')
                    wr = wins / n * 100
                    total_pnl = sum(t['pnl'] for t in trades_list)

                    if wr >= 38 and total_pnl > 0:
                        rec = "YES $100"
                    elif wr >= 30:
                        rec = "DYN $1"
                    else:
                        rec = "OFF"
                else:
                    rec = "N/A"
            else:
                rec = "N/A"
            row += f" {rec:^10} |"
        print(row)

    print()
    print("Legend: YES = Full size | DYN = Dynamic $1 | OFF = Skip | N/A = Not enough data")


if __name__ == "__main__":
    analyze_regimes()
