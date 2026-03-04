# -*- coding: utf-8 -*-
"""Deep analysis of backtest results."""
import openpyxl
from collections import defaultdict
import statistics

FILE = r"G:\BinanceFriend\backtester\output\backtest_signals_part2_GrossPnL-377_WR+36_10020_5643.xlsx"

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb.active

# Collect all data
data = []
for row in range(2, ws.max_row + 1):
    try:
        record = {
            'symbol': ws.cell(row=row, column=3).value,
            'timestamp': ws.cell(row=row, column=4).value,
            'direction': ws.cell(row=row, column=5).value,
            'probability': ws.cell(row=row, column=6).value,
            'confidence': ws.cell(row=row, column=7).value,
            'rr': ws.cell(row=row, column=8).value,
            'filled': ws.cell(row=row, column=23).value,
            'exit_reason': ws.cell(row=row, column=26).value,
            'net_pnl': ws.cell(row=row, column=32).value,
            'net_pct': ws.cell(row=row, column=34).value,
            'hours': ws.cell(row=row, column=35).value,
            'tp1_hit': ws.cell(row=row, column=36).value,
            'tp2_hit': ws.cell(row=row, column=37).value,
            'tp3_hit': ws.cell(row=row, column=38).value,
            'sl_hit': ws.cell(row=row, column=39).value,
            'acc_total': ws.cell(row=row, column=61).value,
            'acc_crowd_bearish': ws.cell(row=row, column=44).value,
            'acc_coordinated_buying': ws.cell(row=row, column=46).value,
            'acc_funding_cheap': ws.cell(row=row, column=42).value,
            'funding_rate_pct': ws.cell(row=row, column=68).value,
            'volume_spike': ws.cell(row=row, column=87).value,
            'trigger_type': ws.cell(row=row, column=101).value,
            'trigger_score': ws.cell(row=row, column=103).value,
            'hour': ws.cell(row=row, column=120).value,
            'day_of_week': ws.cell(row=row, column=122).value,
            'sl_pct': ws.cell(row=row, column=13).value,
        }
        data.append(record)
    except:
        pass

print("=" * 70)
print("ГЛУБОКИЙ АНАЛИЗ БЭКТЕСТА")
print("=" * 70)

# Basic stats
filled = [d for d in data if d['filled'] == 'YES']
not_filled = [d for d in data if d['filled'] == 'NO']

print(f"\n[1] ОБЩАЯ СТАТИСТИКА")
print(f"    Всего сигналов: {len(data)}")
print(f"    Исполнено: {len(filled)} ({len(filled)/len(data)*100:.1f}%)")
print(f"    Не исполнено: {len(not_filled)} ({len(not_filled)/len(data)*100:.1f}%)")

# Win/Loss
wins = [d for d in filled if d['net_pnl'] and float(d['net_pnl']) > 0]
losses = [d for d in filled if d['net_pnl'] and float(d['net_pnl']) < 0]
print(f"    Прибыльных: {len(wins)} ({len(wins)/len(filled)*100:.1f}%)")
print(f"    Убыточных: {len(losses)} ({len(losses)/len(filled)*100:.1f}%)")

# PnL distribution
pnls = [float(d['net_pnl']) for d in filled if d['net_pnl']]
print(f"    Сумма PnL: {sum(pnls):.4f}")
print(f"    Средний PnL: {statistics.mean(pnls):.6f}")
print(f"    Медиана PnL: {statistics.median(pnls):.6f}")

# Exit reasons
print(f"\n[2] ПРИЧИНЫ ВЫХОДА")
exit_stats = defaultdict(lambda: {'count': 0, 'pnl': 0})
for d in filled:
    reason = d['exit_reason'] or 'UNKNOWN'
    exit_stats[reason]['count'] += 1
    if d['net_pnl']:
        exit_stats[reason]['pnl'] += float(d['net_pnl'])

for reason, stats in sorted(exit_stats.items(), key=lambda x: x[1]['pnl']):
    print(f"    {reason:20s}: {stats['count']:5d} сделок, PnL: {stats['pnl']:+.4f}")

# By Symbol
print(f"\n[3] АНАЛИЗ ПО МОНЕТАМ")
symbol_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0, 'pnls': []})
for d in filled:
    sym = d['symbol']
    symbol_stats[sym]['count'] += 1
    if d['net_pnl']:
        pnl = float(d['net_pnl'])
        symbol_stats[sym]['pnl'] += pnl
        symbol_stats[sym]['pnls'].append(pnl)
        if pnl > 0:
            symbol_stats[sym]['wins'] += 1

for sym, stats in sorted(symbol_stats.items(), key=lambda x: x[1]['pnl']):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    avg = statistics.mean(stats['pnls']) if stats['pnls'] else 0
    print(f"    {sym:12s}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+10.4f}, Avg: {avg:+.6f}")

# By Probability
print(f"\n[4] АНАЛИЗ ПО ВЕРОЯТНОСТИ (Probability)")
prob_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    prob = d['probability']
    if prob:
        prob_stats[prob]['count'] += 1
        if d['net_pnl']:
            pnl = float(d['net_pnl'])
            prob_stats[prob]['pnl'] += pnl
            if pnl > 0:
                prob_stats[prob]['wins'] += 1

for prob, stats in sorted(prob_stats.items()):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"    Prob {prob:3.0f}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# By Accumulation Score
print(f"\n[5] АНАЛИЗ ПО ACCUMULATION SCORE")
acc_ranges = [(45, 50), (50, 55), (55, 60), (60, 65), (65, 70), (70, 100)]
for low, high in acc_ranges:
    subset = [d for d in filled if d['acc_total'] and low <= d['acc_total'] < high]
    if subset:
        wins_s = len([d for d in subset if d['net_pnl'] and float(d['net_pnl']) > 0])
        pnl_s = sum(float(d['net_pnl']) for d in subset if d['net_pnl'])
        wr = wins_s/len(subset)*100
        print(f"    Score {low:2d}-{high:2d}: {len(subset):5d} сделок, WR: {wr:5.1f}%, PnL: {pnl_s:+.4f}")

# By Confidence
print(f"\n[6] АНАЛИЗ ПО УВЕРЕННОСТИ (Confidence)")
conf_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    conf = d['confidence'] or 'UNKNOWN'
    conf_stats[conf]['count'] += 1
    if d['net_pnl']:
        pnl = float(d['net_pnl'])
        conf_stats[conf]['pnl'] += pnl
        if pnl > 0:
            conf_stats[conf]['wins'] += 1

for conf, stats in sorted(conf_stats.items(), key=lambda x: x[1]['pnl']):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"    {conf:15s}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# By Hour
print(f"\n[7] АНАЛИЗ ПО ЧАСАМ (UTC)")
hour_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    h = d['hour']
    if h is not None:
        hour_stats[int(h)]['count'] += 1
        if d['net_pnl']:
            pnl = float(d['net_pnl'])
            hour_stats[int(h)]['pnl'] += pnl
            if pnl > 0:
                hour_stats[int(h)]['wins'] += 1

best_hours = sorted(hour_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)[:5]
worst_hours = sorted(hour_stats.items(), key=lambda x: x[1]['pnl'])[:5]
print("    Лучшие часы:")
for h, stats in best_hours:
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"      {h:02d}:00 - {stats['count']:4d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")
print("    Худшие часы:")
for h, stats in worst_hours:
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"      {h:02d}:00 - {stats['count']:4d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# By Day of Week
print(f"\n[8] АНАЛИЗ ПО ДНЯМ НЕДЕЛИ")
dow_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
dow_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    dow = d['day_of_week']
    if dow is not None:
        dow_stats[int(dow)]['count'] += 1
        if d['net_pnl']:
            pnl = float(d['net_pnl'])
            dow_stats[int(dow)]['pnl'] += pnl
            if pnl > 0:
                dow_stats[int(dow)]['wins'] += 1

for dow in range(7):
    stats = dow_stats[dow]
    if stats['count']:
        wr = stats['wins']/stats['count']*100
        print(f"    {dow_names[dow]}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# TP progression
print(f"\n[9] ПРОГРЕССИЯ TP")
tp1_only = len([d for d in filled if d['tp1_hit'] == 'YES' and d['tp2_hit'] != 'YES'])
tp2_only = len([d for d in filled if d['tp2_hit'] == 'YES' and d['tp3_hit'] != 'YES'])
tp3_full = len([d for d in filled if d['tp3_hit'] == 'YES'])
no_tp = len([d for d in filled if d['tp1_hit'] != 'YES'])

print(f"    Без TP (сразу SL/Timeout): {no_tp} ({no_tp/len(filled)*100:.1f}%)")
print(f"    Только TP1: {tp1_only} ({tp1_only/len(filled)*100:.1f}%)")
print(f"    До TP2: {tp2_only} ({tp2_only/len(filled)*100:.1f}%)")
print(f"    Полный TP3: {tp3_full} ({tp3_full/len(filled)*100:.1f}%)")

# Trigger analysis
print(f"\n[10] АНАЛИЗ ПО ТРИГГЕРАМ")
trigger_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    trig = d['trigger_type'] or 'NO_TRIGGER'
    trigger_stats[trig]['count'] += 1
    if d['net_pnl']:
        pnl = float(d['net_pnl'])
        trigger_stats[trig]['pnl'] += pnl
        if pnl > 0:
            trigger_stats[trig]['wins'] += 1

for trig, stats in sorted(trigger_stats.items(), key=lambda x: x[1]['pnl']):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"    {trig:25s}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# Hold time analysis
print(f"\n[11] АНАЛИЗ ВРЕМЕНИ УДЕРЖАНИЯ")
hours_list = [float(d['hours']) for d in filled if d['hours']]
if hours_list:
    print(f"    Среднее время: {statistics.mean(hours_list):.1f} часов")
    print(f"    Медиана: {statistics.median(hours_list):.1f} часов")

    # By outcome
    win_hours = [float(d['hours']) for d in wins if d['hours']]
    loss_hours = [float(d['hours']) for d in losses if d['hours']]
    if win_hours:
        print(f"    Среднее WINS: {statistics.mean(win_hours):.1f} ч (медиана: {statistics.median(win_hours):.1f} ч)")
    if loss_hours:
        print(f"    Среднее LOSSES: {statistics.mean(loss_hours):.1f} ч (медиана: {statistics.median(loss_hours):.1f} ч)")

# SL % analysis
print(f"\n[12] АНАЛИЗ STOP LOSS %")
sl_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    sl = d['sl_pct']
    if sl:
        sl_bucket = round(float(sl))
        sl_stats[sl_bucket]['count'] += 1
        if d['net_pnl']:
            pnl = float(d['net_pnl'])
            sl_stats[sl_bucket]['pnl'] += pnl
            if pnl > 0:
                sl_stats[sl_bucket]['wins'] += 1

for sl, stats in sorted(sl_stats.items()):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"    SL {sl}%: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# Funding analysis
print(f"\n[13] АНАЛИЗ ПО ФАНДИНГУ")
funding_ranges = [(-1, -0.1), (-0.1, -0.01), (-0.01, 0.01), (0.01, 0.1), (0.1, 1)]
for low, high in funding_ranges:
    subset = [d for d in filled if d['funding_rate_pct'] and low <= float(d['funding_rate_pct']) < high]
    if subset:
        wins_s = len([d for d in subset if d['net_pnl'] and float(d['net_pnl']) > 0])
        pnl_s = sum(float(d['net_pnl']) for d in subset if d['net_pnl'])
        wr = wins_s/len(subset)*100
        label = f"{low:+.2f} to {high:+.2f}%"
        print(f"    Funding {label:20s}: {len(subset):5d} сделок, WR: {wr:5.1f}%, PnL: {pnl_s:+.4f}")

# acc_crowd_bearish analysis
print(f"\n[14] АНАЛИЗ acc_crowd_bearish (контрарный сигнал)")
crowd_stats = defaultdict(lambda: {'count': 0, 'wins': 0, 'pnl': 0})
for d in filled:
    crowd = d['acc_crowd_bearish']
    if crowd is not None:
        crowd_stats[int(crowd)]['count'] += 1
        if d['net_pnl']:
            pnl = float(d['net_pnl'])
            crowd_stats[int(crowd)]['pnl'] += pnl
            if pnl > 0:
                crowd_stats[int(crowd)]['wins'] += 1

for crowd, stats in sorted(crowd_stats.items()):
    wr = stats['wins']/stats['count']*100 if stats['count'] else 0
    print(f"    Crowd={crowd:2d}: {stats['count']:5d} сделок, WR: {wr:5.1f}%, PnL: {stats['pnl']:+.4f}")

# Volume spike analysis
print(f"\n[15] АНАЛИЗ VOLUME SPIKE")
vol_ranges = [(0, 0.5), (0.5, 0.8), (0.8, 1.0), (1.0, 1.2), (1.2, 1.5), (1.5, 2.0), (2.0, 10)]
for low, high in vol_ranges:
    subset = [d for d in filled if d['volume_spike'] and low <= float(d['volume_spike']) < high]
    if subset:
        wins_s = len([d for d in subset if d['net_pnl'] and float(d['net_pnl']) > 0])
        pnl_s = sum(float(d['net_pnl']) for d in subset if d['net_pnl'])
        wr = wins_s/len(subset)*100
        print(f"    Spike {low:.1f}-{high:.1f}: {len(subset):5d} сделок, WR: {wr:5.1f}%, PnL: {pnl_s:+.4f}")

print("\n" + "=" * 70)
