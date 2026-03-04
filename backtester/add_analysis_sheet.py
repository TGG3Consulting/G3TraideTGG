# -*- coding: utf-8 -*-
"""
Add analysis sheet to backtest Excel file.
Creates summary tables by Symbol and by Period with Excel formulas.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
from datetime import datetime

INPUT_FILE = r"G:\BinanceFriend\backtester\output\backtest_signals_part10__GrossPnL+0.4_WR+35.5_2202_1377.xlsx"

def main():
    print(f"Loading: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # Get data from main sheet
    ws_main = wb["Backtest Results"]
    max_row = ws_main.max_row

    print(f"Data rows: {max_row - 1}")

    # Collect unique symbols and periods
    symbols = set()
    periods = set()  # YYYY-MM format

    # Also collect raw data for period-symbol breakdown
    period_symbol_data = defaultdict(lambda: defaultdict(float))

    for row in range(2, max_row + 1):
        symbol = ws_main.cell(row=row, column=3).value  # Column C
        timestamp = ws_main.cell(row=row, column=4).value  # Column D
        net_pnl = ws_main.cell(row=row, column=32).value  # Column AF

        if symbol:
            symbols.add(symbol)

        if timestamp:
            # Parse timestamp (format: "2025-09-03 17:40")
            try:
                if isinstance(timestamp, str):
                    dt = datetime.strptime(timestamp[:7], "%Y-%m")
                    period = timestamp[:7]  # YYYY-MM
                elif isinstance(timestamp, datetime):
                    period = timestamp.strftime("%Y-%m")
                else:
                    continue
                periods.add(period)

                # Accumulate for period-symbol breakdown
                if net_pnl and symbol:
                    try:
                        period_symbol_data[period][symbol] += float(net_pnl)
                    except (ValueError, TypeError):
                        pass
            except:
                pass

    symbols = sorted(symbols)
    periods = sorted(periods)

    print(f"Symbols: {len(symbols)}")
    print(f"Periods: {len(periods)}")

    # Create Analysis sheet
    if "Analysis" in wb.sheetnames:
        del wb["Analysis"]

    ws = wb.create_sheet("Analysis")

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # =============================================
    # TABLE 1: PnL by Symbol (sorted by Net PnL)
    # =============================================
    ws.cell(row=1, column=1, value="PnL BY SYMBOL (sorted: worst to best)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Headers
    headers1 = ["Symbol", "Total Net PnL", "Trades Count", "Avg PnL per Trade", "Win Rate %"]
    for col, header in enumerate(headers1, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Calculate data for each symbol using Python (more reliable than SUMIF for large files)
    symbol_stats = {}
    for symbol in symbols:
        total_pnl = 0
        count = 0
        wins = 0

        for row in range(2, max_row + 1):
            sym = ws_main.cell(row=row, column=3).value
            if sym == symbol:
                net_pnl = ws_main.cell(row=row, column=32).value
                if net_pnl is not None:
                    try:
                        pnl_val = float(net_pnl)
                        total_pnl += pnl_val
                        count += 1
                        if pnl_val > 0:
                            wins += 1
                    except:
                        pass

        avg_pnl = total_pnl / count if count > 0 else 0
        win_rate = (wins / count * 100) if count > 0 else 0
        symbol_stats[symbol] = {
            'total_pnl': total_pnl,
            'count': count,
            'avg_pnl': avg_pnl,
            'win_rate': win_rate
        }

    # Sort by total PnL (worst to best)
    sorted_symbols = sorted(symbols, key=lambda s: symbol_stats[s]['total_pnl'])

    # Write data
    for i, symbol in enumerate(sorted_symbols, 3):
        stats = symbol_stats[symbol]

        ws.cell(row=i, column=1, value=symbol).border = border

        cell_pnl = ws.cell(row=i, column=2, value=round(stats['total_pnl'], 8))
        cell_pnl.border = border
        cell_pnl.number_format = '0.00000000'
        if stats['total_pnl'] > 0:
            cell_pnl.fill = green_fill
        elif stats['total_pnl'] < 0:
            cell_pnl.fill = red_fill

        ws.cell(row=i, column=3, value=stats['count']).border = border

        cell_avg = ws.cell(row=i, column=4, value=round(stats['avg_pnl'], 8))
        cell_avg.border = border
        cell_avg.number_format = '0.00000000'

        cell_wr = ws.cell(row=i, column=5, value=round(stats['win_rate'], 2))
        cell_wr.border = border
        cell_wr.number_format = '0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12

    # =============================================
    # TABLE 2: PnL by Period (sorted by Net PnL)
    # =============================================
    start_row2 = len(sorted_symbols) + 5

    ws.cell(row=start_row2, column=1, value="PnL BY PERIOD (sorted: worst to best)")
    ws.cell(row=start_row2, column=1).font = Font(bold=True, size=14)

    headers2 = ["Period", "Total Net PnL", "Trades Count", "Best Symbol", "Best PnL", "Worst Symbol", "Worst PnL"]
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=start_row2 + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Calculate data for each period
    period_stats = {}
    for period in periods:
        total_pnl = 0
        count = 0

        for row in range(2, max_row + 1):
            timestamp = ws_main.cell(row=row, column=4).value
            if timestamp:
                try:
                    if isinstance(timestamp, str):
                        ts_period = timestamp[:7]
                    elif isinstance(timestamp, datetime):
                        ts_period = timestamp.strftime("%Y-%m")
                    else:
                        continue

                    if ts_period == period:
                        net_pnl = ws_main.cell(row=row, column=32).value
                        if net_pnl is not None:
                            try:
                                total_pnl += float(net_pnl)
                                count += 1
                            except:
                                pass
                except:
                    pass

        # Find best and worst symbol for this period
        sym_pnls = period_symbol_data[period]
        if sym_pnls:
            best_sym = max(sym_pnls.keys(), key=lambda s: sym_pnls[s])
            worst_sym = min(sym_pnls.keys(), key=lambda s: sym_pnls[s])
        else:
            best_sym = worst_sym = "-"

        period_stats[period] = {
            'total_pnl': total_pnl,
            'count': count,
            'best_sym': best_sym,
            'best_pnl': sym_pnls.get(best_sym, 0) if sym_pnls else 0,
            'worst_sym': worst_sym,
            'worst_pnl': sym_pnls.get(worst_sym, 0) if sym_pnls else 0,
        }

    # Sort by total PnL (worst to best)
    sorted_periods = sorted(periods, key=lambda p: period_stats[p]['total_pnl'])

    # Write data
    for i, period in enumerate(sorted_periods, start_row2 + 2):
        stats = period_stats[period]

        ws.cell(row=i, column=1, value=period).border = border

        cell_pnl = ws.cell(row=i, column=2, value=round(stats['total_pnl'], 8))
        cell_pnl.border = border
        cell_pnl.number_format = '0.00000000'
        if stats['total_pnl'] > 0:
            cell_pnl.fill = green_fill
        elif stats['total_pnl'] < 0:
            cell_pnl.fill = red_fill

        ws.cell(row=i, column=3, value=stats['count']).border = border
        ws.cell(row=i, column=4, value=stats['best_sym']).border = border

        cell_best = ws.cell(row=i, column=5, value=round(stats['best_pnl'], 8))
        cell_best.border = border
        cell_best.number_format = '0.00000000'
        if stats['best_pnl'] > 0:
            cell_best.fill = green_fill

        ws.cell(row=i, column=6, value=stats['worst_sym']).border = border

        cell_worst = ws.cell(row=i, column=7, value=round(stats['worst_pnl'], 8))
        cell_worst.border = border
        cell_worst.number_format = '0.00000000'
        if stats['worst_pnl'] < 0:
            cell_worst.fill = red_fill

    # Column widths for table 2
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18

    # =============================================
    # TABLE 3: Period-Symbol Matrix (detailed breakdown)
    # =============================================
    start_row3 = start_row2 + len(sorted_periods) + 5

    ws.cell(row=start_row3, column=1, value="PERIOD-SYMBOL MATRIX (Net PnL)")
    ws.cell(row=start_row3, column=1).font = Font(bold=True, size=14)

    # Headers: Period | Symbol1 | Symbol2 | ... | Total
    ws.cell(row=start_row3 + 1, column=1, value="Period").font = header_font
    ws.cell(row=start_row3 + 1, column=1).fill = header_fill
    ws.cell(row=start_row3 + 1, column=1).border = border

    for col, symbol in enumerate(sorted_symbols, 2):
        cell = ws.cell(row=start_row3 + 1, column=col, value=symbol)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    total_col = len(sorted_symbols) + 2
    ws.cell(row=start_row3 + 1, column=total_col, value="TOTAL").font = header_font
    ws.cell(row=start_row3 + 1, column=total_col).fill = header_fill
    ws.cell(row=start_row3 + 1, column=total_col).border = border

    # Data rows
    for row_idx, period in enumerate(sorted_periods, start_row3 + 2):
        ws.cell(row=row_idx, column=1, value=period).border = border

        row_total = 0
        for col, symbol in enumerate(sorted_symbols, 2):
            pnl = period_symbol_data[period].get(symbol, 0)
            cell = ws.cell(row=row_idx, column=col, value=round(pnl, 8) if pnl != 0 else "")
            cell.border = border
            cell.number_format = '0.00000000'
            if pnl > 0:
                cell.fill = green_fill
            elif pnl < 0:
                cell.fill = red_fill
            row_total += pnl

        # Total column
        cell_total = ws.cell(row=row_idx, column=total_col, value=round(row_total, 8))
        cell_total.border = border
        cell_total.number_format = '0.00000000'
        cell_total.font = Font(bold=True)
        if row_total > 0:
            cell_total.fill = green_fill
        elif row_total < 0:
            cell_total.fill = red_fill

    # Symbol totals row
    totals_row = start_row3 + 2 + len(sorted_periods)
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=totals_row, column=1).fill = header_fill
    ws.cell(row=totals_row, column=1).border = border

    grand_total = 0
    for col, symbol in enumerate(sorted_symbols, 2):
        sym_total = sum(period_symbol_data[p].get(symbol, 0) for p in periods)
        cell = ws.cell(row=totals_row, column=col, value=round(sym_total, 8))
        cell.border = border
        cell.font = Font(bold=True)
        cell.number_format = '0.00000000'
        if sym_total > 0:
            cell.fill = green_fill
        elif sym_total < 0:
            cell.fill = red_fill
        grand_total += sym_total

    cell_grand = ws.cell(row=totals_row, column=total_col, value=round(grand_total, 8))
    cell_grand.border = border
    cell_grand.font = Font(bold=True, size=12)
    cell_grand.number_format = '0.00000000'

    # Save
    print(f"\nSaving...")
    wb.save(INPUT_FILE)
    print(f"Done! Added 'Analysis' sheet with 3 tables.")

if __name__ == "__main__":
    main()
